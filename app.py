"""
BIM Infra Solutions — Custom CRM (Flask Web App)
Run: python app.py
Open: http://localhost:5000
"""
import os
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
import database as db
from zoho_mail import mail_client, SEQUENCE_SCHEDULE

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "bim-infra-crm-2025")

# ── LOGIN SETUP ────────────────────────────────────────────────────────────────

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Please log in to access the CRM."
login_manager.login_message_category = "warning"

# Multi-user table — add/remove users here
# All default to password: Bim@2025  (change per-user in .env or regenerate hash)
CRM_USERS = {
    "kishan": {
        "id": "1",
        "display": "Kishan",
        "hash": os.getenv("CRM_HASH_KISHAN",
            "scrypt:32768:8:1$DPjW6lPlfNm2wDBp$5740c2e521ec2a8d728ab9ec15cc626ce78e9b104d9705a3a4e0f07ee05f56f155729fe213e76434d0fc491fea364a5a780fb571516a683f585d4a0be4a737d6"),
    },
    "hirakraj": {
        "id": "2",
        "display": "Hirakraj",
        "hash": os.getenv("CRM_HASH_HIRAKRAJ",
            "scrypt:32768:8:1$tA35OeKwmVXp9icD$ed300560f7ec1439a40377ceec8c09b79cf1cea262a8d990fdb208768c4ac45bf937477c6f351744794d956257a53a9655c6a297f0ecff958568ae524a00abc2"),
    },
    "tirth": {
        "id": "3",
        "display": "Tirth",
        "hash": os.getenv("CRM_HASH_TIRTH",
            "scrypt:32768:8:1$YWy1VzOHFotCk5Zc$29c869fd5761b1a5006d3e5e3067b0ac1d9f914c4613f9274c0bd8188d32e3fcde1b6935b3fe9e078e52861569a05062ca1fc34967f95d18f2e403671a89121c"),
    },
    "jenish": {
        "id": "4",
        "display": "Jenish",
        "hash": os.getenv("CRM_HASH_JENISH",
            "scrypt:32768:8:1$U6YIKHykWKLffUN8$e15fb6578f7e96048a410d43a4a9a5a6b3b2570a1d549f3ac98b653cc04695b67aa724349cd4167c055b791071e4601244b7e3ac6c7afde83374680e6c7996ba"),
    },
}

# Reverse lookup by user ID
_ID_TO_USER = {v["id"]: (k, v) for k, v in CRM_USERS.items()}


class CRMUser(UserMixin):
    def __init__(self, uid, username, display):
        self.id       = uid
        self.username = username
        self.display  = display


@login_manager.user_loader
def load_user(user_id):
    entry = _ID_TO_USER.get(user_id)
    if entry:
        username, info = entry
        return CRMUser(info["id"], username, info["display"])
    return None


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        info = CRM_USERS.get(username)
        if info and check_password_hash(info["hash"], password):
            user = CRMUser(info["id"], username, info["display"])
            login_user(user, remember=request.form.get("remember") == "on")
            next_page = request.args.get("next")
            return redirect(next_page or url_for("dashboard"))
        flash("Invalid username or password.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_pw  = request.form.get("current_password", "")
        new_pw      = request.form.get("new_password", "")
        confirm_pw  = request.form.get("confirm_password", "")

        username = current_user.username
        info     = CRM_USERS.get(username)

        # Validate current password
        if not check_password_hash(info["hash"], current_pw):
            flash("Current password is incorrect.", "danger")
            return render_template("change_password.html")

        if len(new_pw) < 6:
            flash("New password must be at least 6 characters.", "danger")
            return render_template("change_password.html")

        if new_pw != confirm_pw:
            flash("New passwords do not match.", "danger")
            return render_template("change_password.html")

        # Generate new hash
        new_hash = generate_password_hash(new_pw)

        # Update in-memory
        CRM_USERS[username]["hash"] = new_hash
        # Rebuild reverse lookup
        global _ID_TO_USER
        _ID_TO_USER = {v["id"]: (k, v) for k, v in CRM_USERS.items()}

        # Persist to .env so it survives restart
        env_key  = f"CRM_HASH_{username.upper()}"
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
        _update_env(env_path, env_key, new_hash)

        flash("Password changed successfully!", "success")
        return redirect(url_for("dashboard"))

    return render_template("change_password.html")


def _update_env(env_path: str, key: str, value: str):
    """Write or update a key=value line in the .env file."""
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
    else:
        lines = []

    found = False
    new_lines = []
    for line in lines:
        if line.startswith(f"{key}="):
            new_lines.append(f"{key}={value}\n")
            found = True
        else:
            new_lines.append(line)

    if not found:
        # Add a blank line before if file doesn't end with one
        if new_lines and not new_lines[-1].endswith("\n"):
            new_lines.append("\n")
        new_lines.append(f"{key}={value}\n")

    with open(env_path, "w", encoding="utf-8") as f:
        f.writelines(new_lines)

# ──────────────────────────────────────────────────────────────────────────────

db.init_db()


# ── AUTH GUARD — protect every route except login ─────────────────────────────

@app.before_request
def require_login():
    public = {"login", "static"}
    if request.endpoint and request.endpoint not in public and not current_user.is_authenticated:
        return redirect(url_for("login", next=request.url))


# ── DASHBOARD ──────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    stats           = db.get_stats()
    tasks           = db.get_tasks(status="Not Started")
    due_leads       = db.get_due_email_leads()
    invalid_leads   = db.get_invalid_leads_with_bounce()
    my_tasks        = db.get_team_tasks(assigned_to=current_user.username, status=None)
    my_tasks        = [t for t in my_tasks if t.get("status") not in ("Done", "Cancelled")]
    team_tasks      = db.get_team_tasks(assigned_to=None, status=None)
    team_tasks      = [t for t in team_tasks if t.get("status") not in ("Done", "Cancelled")]
    my_resp         = db.get_responsibilities(assigned_to=current_user.username)
    my_resp         = [r for r in my_resp if r.get("status") == "Active"]
    recent_invoices = db.get_invoices()[:5]
    inv_summary     = db.get_invoice_summary()
    projects        = db.get_projects(status="Active")
    return render_template("dashboard.html", stats=stats, tasks=tasks,
                           due_leads=due_leads, invalid_leads=invalid_leads,
                           my_tasks=my_tasks, team_tasks=team_tasks,
                           my_resp=my_resp, recent_invoices=recent_invoices,
                           inv_summary=inv_summary, projects=projects,
                           now=datetime.now())


# ── LEADS LIST ─────────────────────────────────────────────────────────────────

@app.route("/leads")
def leads():
    search   = request.args.get("search", "")
    status   = request.args.get("status", "")
    country  = request.args.get("country", "")
    template = request.args.get("template", "")
    all_leads = db.get_all_leads(
        search=search or None,
        status=status or None,
        country=country or None,
        template=template or None,
    )
    return render_template("leads.html", leads=all_leads,
                           search=search, status=status,
                           country=country, template=template)


# ── ADD LEAD ───────────────────────────────────────────────────────────────────

@app.route("/leads/new", methods=["GET", "POST"])
def new_lead():
    if request.method == "POST":
        data = _form_to_lead(request.form)
        data["status"] = "New"
        try:
            lead_id = db.create_lead(data)
            flash("Lead created successfully!", "success")
            return redirect(url_for("lead_detail", lead_id=lead_id))
        except Exception as e:
            flash(f"Error creating lead: {str(e)}", "danger")
    return render_template("lead_form.html", lead=None, action="New")


# ── LEAD DETAIL ────────────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>")
def lead_detail(lead_id):
    lead = db.get_lead(lead_id)
    if not lead:
        flash("Lead not found.", "danger")
        return redirect(url_for("leads"))

    email_logs = db.get_email_logs(lead_id)
    tasks      = db.get_tasks(lead_id=lead_id)

    template  = lead.get("email_template", "A")
    step      = lead.get("email_sequence_step", 0)
    days      = SEQUENCE_SCHEDULE.get(template, [0, 4, 9])

    can_send_next  = False
    next_send_date = None

    if step < 3:
        if step == 0:
            can_send_next = True
        elif lead.get("last_email_sent"):
            last_sent    = datetime.fromisoformat(str(lead["last_email_sent"]))
            days_since   = (datetime.utcnow() - last_sent).days
            required     = days[step] - days[step - 1]
            can_send_next = days_since >= required
            if not can_send_next:
                next_send_date = (last_sent + timedelta(days=required)).strftime("%b %d, %Y")

    wa_logs = db.get_whatsapp_logs(lead_id)
    return render_template("lead_detail.html",
                           lead=lead,
                           email_logs=email_logs,
                           tasks=tasks,
                           can_send_next=can_send_next,
                           next_send_date=next_send_date,
                           sequence_step=step,
                           wa_logs=wa_logs,
                           wa_templates=WA_TEMPLATES,
                           twilio_enabled=bool(os.getenv("TWILIO_ACCOUNT_SID")))


# ── EDIT LEAD ──────────────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>/edit", methods=["GET", "POST"])
def edit_lead(lead_id):
    lead = db.get_lead(lead_id)
    if not lead:
        flash("Lead not found.", "danger")
        return redirect(url_for("leads"))

    if request.method == "POST":
        data = _form_to_lead(request.form)
        data["status"] = request.form.get("status", lead["status"])
        db.update_lead(lead_id, data)
        flash("Lead updated!", "success")
        return redirect(url_for("lead_detail", lead_id=lead_id))

    return render_template("lead_form.html", lead=lead, action="Edit")


# ── DELETE LEAD ────────────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>/delete", methods=["POST"])
def delete_lead(lead_id):
    db.delete_lead(lead_id)
    flash("Lead deleted.", "info")
    return redirect(url_for("leads"))


# ── UPDATE STATUS ──────────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>/status", methods=["POST"])
def update_status(lead_id):
    status = request.form.get("status", "New")
    db.update_lead_status(lead_id, status)
    flash(f"Status updated to {status}.", "success")
    return redirect(url_for("lead_detail", lead_id=lead_id))


# ── SEND EMAIL ─────────────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>/send-email", methods=["POST"])
def send_email(lead_id):
    lead = db.get_lead(lead_id)
    if not lead:
        flash("Lead not found.", "danger")
        return redirect(url_for("leads"))

    step = lead.get("email_sequence_step", 0)
    if step >= 3:
        flash("All 3 sequence emails already sent for this lead.", "warning")
        return redirect(url_for("lead_detail", lead_id=lead_id))

    try:
        user_cfg      = db.get_user_settings(current_user.username)
        custom_tpl    = db.get_email_template(lead.get("email_template", "A"), step)
        success, subject, body = mail_client.send_sequence_email(
            lead, step,
            user_settings=user_cfg or None,
            custom_template=custom_tpl or None,
        )
    except Exception as e:
        err = str(e)
        if "invalid_code" in err or "invalid_token" in err:
            flash("Zoho token expired — run get_token.py to refresh your credentials, then restart the app.", "danger")
        else:
            flash(f"Email error: {err}", "danger")
        return redirect(url_for("lead_detail", lead_id=lead_id))

    if success:
        db.log_email(lead_id, subject, body, lead.get("email_template", "A"), step)
        db.advance_sequence_step(lead_id)
        if lead.get("status") == "New":
            db.update_lead_status(lead_id, "Contacted")
        flash(f"Email {step + 1}/3 sent to {lead['email']}!", "success")
    else:
        flash("Email failed. Run get_token.py to refresh Zoho credentials.", "danger")

    return redirect(url_for("lead_detail", lead_id=lead_id))


# ── MARK EMAIL OPENED ──────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>/mark-opened", methods=["POST"])
def mark_opened(lead_id):
    email_log_id = int(request.form.get("email_log_id", 0))
    db.mark_email_opened(email_log_id)

    # Check if lead qualifies as hot (2+ opens)
    lead = db.get_lead(lead_id)
    logs = db.get_email_logs(lead_id)
    total_opens = sum(l.get("open_count", 0) for l in logs)

    if total_opens >= 2 and lead.get("status") != "Hot":
        db.update_lead_status(lead_id, "Hot", "Immediate Follow-up")
        db.create_task(
            lead_id,
            f"HOT LEAD — Call {lead['first_name']} {lead.get('last_name', '')} ASAP",
            (datetime.utcnow() + timedelta(hours=2)).isoformat(),
            priority="High",
            description="Auto-generated: 2+ email opens detected.",
        )
        flash("Lead marked HOT — urgent task created!", "warning")
    else:
        flash("Email marked as opened.", "success")

    return redirect(url_for("lead_detail", lead_id=lead_id))


# ── MARK EMAIL CLICKED ─────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>/mark-bounced", methods=["POST"])
def mark_bounced(lead_id):
    email_log_id = int(request.form.get("email_log_id", 0))
    reason       = request.form.get("reason", "Email bounced / address not found")
    db.mark_email_bounced(email_log_id, reason)

    lead = db.get_lead(lead_id)
    # Rewind sequence step so next email can be retried with corrected address
    db.update_lead(lead_id, {"email_sequence_step": max(0, lead.get("email_sequence_step", 1) - 1)})
    db.update_lead_status(lead_id, "Invalid")
    flash(
        f"Email marked as FAILED/BOUNCED for {lead['email']}. "
        f"Update the email address and resend. Status set to Invalid.",
        "warning",
    )
    return redirect(url_for("lead_detail", lead_id=lead_id))


@app.route("/leads/<int:lead_id>/mark-clicked", methods=["POST"])
def mark_clicked(lead_id):
    email_log_id = int(request.form.get("email_log_id", 0))
    db.mark_email_clicked(email_log_id)

    lead = db.get_lead(lead_id)
    if lead.get("status") not in ("Hot", "Engaged"):
        db.update_lead_status(lead_id, "Hot", "Link Clicked — Follow Up")
        db.create_task(
            lead_id,
            f"HOT LEAD (link clicked) — Follow up with {lead['first_name']} within 2 hours",
            (datetime.utcnow() + timedelta(hours=2)).isoformat(),
            priority="High",
        )
        flash("Lead marked HOT — link click detected!", "warning")
    else:
        flash("Email marked as clicked.", "success")

    return redirect(url_for("lead_detail", lead_id=lead_id))


# ── TASKS ──────────────────────────────────────────────────────────────────────

@app.route("/leads/<int:lead_id>/task", methods=["POST"])
def create_task(lead_id):
    subject     = request.form.get("subject", "")
    due_date    = request.form.get("due_date", "")
    priority    = request.form.get("priority", "Medium")
    description = request.form.get("description", "")
    if subject:
        db.create_task(lead_id, subject, due_date, priority, description)
        flash("Task created!", "success")
    return redirect(url_for("lead_detail", lead_id=lead_id))


@app.route("/tasks/<int:task_id>/complete", methods=["POST"])
def complete_task(task_id):
    db.complete_task(task_id)
    flash("Task marked complete.", "success")
    return redirect(request.referrer or url_for("dashboard"))


# ── BULK SEND EMAILS ──────────────────────────────────────────────────────────

@app.route("/leads/send-all", methods=["POST"])
@login_required
def send_all_emails():
    all_leads   = db.get_all_leads()
    user_cfg    = db.get_user_settings(current_user.username)

    sent         = 0
    skip_done    = 0   # sequence complete (step >= 3)
    skip_status  = 0   # Invalid / Unsubscribed
    skip_timing  = 0   # too early
    skip_failed  = 0   # Zoho send returned False
    errors       = []

    for lead in all_leads:
        step   = lead.get("email_sequence_step", 0)
        status = lead.get("status", "")

        # Skip leads that finished sequence
        if step >= 3:
            skip_done += 1
            continue

        # Skip unsubscribed / invalid
        if status in ("Unsubscribed", "Invalid"):
            skip_status += 1
            continue

        # Check sequence timing
        if step > 0 and lead.get("last_email_sent"):
            template     = lead.get("email_template", "A")
            days         = SEQUENCE_SCHEDULE.get(template, [0, 4, 9])
            last_sent    = datetime.fromisoformat(str(lead["last_email_sent"]))
            days_since   = (datetime.utcnow() - last_sent).days
            required     = days[step] - days[step - 1]
            if days_since < required:
                skip_timing += 1
                continue

        # Send the email
        try:
            custom_tpl = db.get_email_template(lead.get("email_template", "A"), step)
            success, subject, body = mail_client.send_sequence_email(
                lead, step,
                user_settings=user_cfg or None,
                custom_template=custom_tpl or None,
            )
        except Exception as e:
            skip_failed += 1
            err_msg = str(e)
            if len(errors) < 3:
                errors.append(f"{lead['email']}: {err_msg[:80]}")
            continue

        if success:
            db.log_email(lead["id"], subject, body, lead.get("email_template", "A"), step)
            db.advance_sequence_step(lead["id"])
            if lead.get("status") == "New":
                db.update_lead_status(lead["id"], "Contacted")
            sent += 1
        else:
            skip_failed += 1

    parts = [f"<strong>{sent} sent</strong>"]
    if skip_done:    parts.append(f"{skip_done} sequence complete")
    if skip_status:  parts.append(f"{skip_status} invalid/unsub")
    if skip_timing:  parts.append(f"{skip_timing} too early")
    if skip_failed:  parts.append(f"<strong style='color:#c62828'>{skip_failed} failed (check Zoho credentials)</strong>")

    msg = "Bulk send: " + " | ".join(parts)
    if errors:
        msg += "<br><small style='color:#c62828'>Errors: " + "; ".join(errors) + "</small>"

    flash(msg, "success" if sent > 0 else ("warning" if skip_failed == 0 else "danger"))
    return redirect(url_for("leads"))


# ── IMPORT FROM GOOGLE SHEET ──────────────────────────────────────────────────

@app.route("/leads/import-sheet", methods=["GET", "POST"])
def import_from_sheet():
    if request.method == "POST":
        import csv, io, requests as req
        sheet_id = request.form.get("sheet_id", "1buAYF8WykRsNToMw8UjAHct_EOX8TyH8")
        template = request.form.get("template", "D")
        try:
            r       = req.get(f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv", timeout=15)
            content = r.content.decode("utf-8", errors="replace")
            rows    = list(csv.reader(io.StringIO(content)))
            imported, skipped = 0, 0
            for row in rows[3:]:
                if not row or not row[0].strip() or not row[0].strip().isdigit():
                    continue
                email   = row[6].strip() if len(row) > 6 else ""
                company = row[1].strip() if len(row) > 1 else ""
                name    = row[3].strip() if len(row) > 3 else ""
                parts   = name.split() if name else []
                if not email or email.lower() in ("", "nan"):
                    skipped += 1
                    continue
                cls      = row[2].strip() if len(row) > 2 else ""
                priority = 90 if cls == "A+" else 75 if cls == "A" else 60 if cls == "B" else 45
                data = {
                    "first_name"            : parts[0] if parts else company.split()[0],
                    "last_name"             : " ".join(parts[1:]) if len(parts) > 1 else "",
                    "email"                 : email,
                    "company"               : company,
                    "title"                 : row[4].strip() if len(row) > 4 else "",
                    "phone"                 : row[5].strip() if len(row) > 5 else "",
                    "website"               : row[9].strip() if len(row) > 9 else "",
                    "city"                  : row[8].strip() if len(row) > 8 else "",
                    "country"               : "India",
                    "industry"              : "Geospatial / Drone / Survey",
                    "status"                : "New",
                    "priority_score"        : priority,
                    "services_needed"       : row[10].strip() if len(row) > 10 else "",
                    "outsourcing_likelihood": "High" if cls in ("A+", "A") else "Medium",
                    "pitch_angle"           : "INFRA X Drone Progress Monitoring",
                    "email_template"        : template,
                    "linkedin_url"          : "",
                    "follow_up_stage"       : "",
                    "description"           : f"Survey of India Empanelled — Class {cls}",
                }
                try:
                    db.create_lead(data)
                    imported += 1
                except Exception:
                    skipped += 1
            flash(f"Sheet import complete: {imported} imported, {skipped} skipped.", "success")
        except Exception as e:
            flash(f"Import failed: {str(e)}", "danger")
        return redirect(url_for("leads"))
    return render_template("import_sheet.html")


# ── PIPELINE (KANBAN) ─────────────────────────────────────────────────────────

@app.route("/pipeline")
def pipeline():
    all_leads = db.get_all_leads()
    stages = ["New", "Contacted", "Warm", "Hot", "Engaged", "Cold", "Unsubscribed", "Invalid"]
    stage_colors = {
        "New":          "#1B3A6B",
        "Contacted":    "#1565c0",
        "Warm":         "#e65100",
        "Hot":          "#b71c1c",
        "Engaged":      "#4a148c",
        "Cold":         "#283593",
        "Unsubscribed": "#616161",
        "Invalid":      "#c62828",
    }
    columns = {}
    totals  = {}
    for s in stages:
        cols_leads = [l for l in all_leads if l["status"] == s]
        columns[s] = cols_leads
        totals[s]  = len(cols_leads)
    return render_template("pipeline.html",
                           stages=stages,
                           columns=columns,
                           totals=totals,
                           stage_colors=stage_colors)


@app.route("/leads/<int:lead_id>/move-stage", methods=["POST"])
def move_stage(lead_id):
    new_status = request.form.get("status", "New")
    db.update_lead_status(lead_id, new_status)
    return ("", 204)


# ── REPLIES INBOX ─────────────────────────────────────────────────────────────

@app.route("/replies")
def replies():
    priority = request.args.get("priority", "")
    all_replies = db.get_replies(priority=priority or None, status=None)
    counts      = db.reply_counts()
    return render_template("replies.html",
                           replies=all_replies,
                           counts=counts,
                           active_priority=priority)


@app.route("/replies/sync-delivery")
def sync_delivery():
    """Scan Zoho inbox for bounces, OOO, and replies. Auto-mark leads."""
    all_leads  = db.get_all_leads()
    lead_map   = {l["email"].lower(): l for l in all_leads}
    lead_emails = list(lead_map.keys())

    try:
        status = mail_client.fetch_delivery_status(lead_emails)
    except Exception as e:
        flash(f"Zoho scan failed: {e}", "danger")
        return redirect(url_for("replies"))

    bounced_count = 0
    replied_count = 0
    ooo_count     = 0

    delayed_count = 0

    # ── Handle bounces / delivery failures ───────────────────────────
    for b in status.get("bounced", []):
        email = b.get("email")
        if not email:
            continue
        lead = lead_map.get(email)
        if not lead:
            continue

        is_delayed   = b.get("is_delayed", False)
        is_permanent = b.get("is_permanent", True)
        reason       = b.get("reason", "Delivery failure")
        raw_sender   = b.get("raw_sender", "")
        subject_line = b.get("subject", "")

        if is_delayed and not is_permanent:
            # Temporary failure (421, 450) — log as warning, don't invalidate lead
            db.add_reply(
                lead_id    = lead["id"],
                from_email = raw_sender,
                subject    = f"⏳ DELAYED: {subject_line}",
                body       = reason,
                priority   = "Medium",
                source     = "Zoho Auto-Sync",
            )
            delayed_count += 1
        else:
            # Permanent failure — mark email bounced, set lead Invalid
            logs = db.get_email_logs(lead["id"])
            if logs:
                db.mark_email_bounced(logs[0]["id"], reason)
            db.update_lead_status(lead["id"], "Invalid")
            db.add_reply(
                lead_id    = lead["id"],
                from_email = raw_sender,
                subject    = f"❌ BOUNCE: {subject_line}",
                body       = reason,
                priority   = "High",
                source     = "Zoho Auto-Sync",
            )
            bounced_count += 1

    # ── Handle real replies ───────────────────────────────────────────
    for rp in status.get("replied", []):
        lead = lead_map.get(rp["email"])
        if not lead:
            continue
        score = lead.get("priority_score", 0)
        auto_priority = "High" if score >= 80 else "Medium" if score >= 50 else "Low"
        db.add_reply(
            lead_id    = lead["id"],
            from_email = rp["email"],
            subject    = rp["subject"],
            body       = rp["body"],
            priority   = auto_priority,
            source     = "Zoho Auto-Sync",
        )
        db.update_lead_status(lead["id"], "Engaged")
        db.create_task(
            lead["id"],
            f"Reply received — respond to {lead['first_name']} within 1 hour",
            (datetime.utcnow() + timedelta(hours=1)).isoformat(),
            priority="High",
        )
        replied_count += 1

    # ── Handle OOO ────────────────────────────────────────────────────
    for o in status.get("ooo", []):
        lead = lead_map.get(o["email"])
        if not lead:
            continue
        db.add_reply(
            lead_id    = lead["id"],
            from_email = o["email"],
            subject    = f"OOO: {o['subject']}",
            body       = "Out of office auto-reply received.",
            priority   = "Low",
            source     = "Zoho Auto-Sync",
        )
        db.update_lead_status(lead["id"], "Warm")
        ooo_count += 1

    total = bounced_count + delayed_count + replied_count + ooo_count
    parts = []
    if bounced_count:  parts.append(f"{bounced_count} hard bounce(s)")
    if delayed_count:  parts.append(f"{delayed_count} delayed (temporary)")
    if replied_count:  parts.append(f"{replied_count} reply(ies)")
    if ooo_count:      parts.append(f"{ooo_count} out-of-office")
    msg = "Zoho scan complete — " + (", ".join(parts) if parts else "nothing new found")
    flash(msg, "success" if total > 0 else "info")
    return redirect(url_for("replies"))


@app.route("/replies/sync")
def sync_replies():
    """Pull replies from Zoho Mail inbox."""
    all_leads   = db.get_all_leads()
    lead_map    = {l["email"].lower(): l for l in all_leads}
    lead_emails = list(lead_map.keys())

    fetched = mail_client.fetch_inbox_replies(lead_emails)
    added   = 0

    for msg in fetched:
        lead = lead_map.get(msg["from_email"].lower())
        if lead:
            # Auto-priority based on lead score
            score = lead.get("priority_score", 0)
            auto_priority = "High" if score >= 80 else "Medium" if score >= 50 else "Low"

            db.add_reply(
                lead_id    = lead["id"],
                from_email = msg["from_email"],
                subject    = msg["subject"],
                body       = msg["body"],
                priority   = auto_priority,
                source     = "Zoho Mail Sync",
            )
            db.update_lead_status(lead["id"], "Engaged")
            db.create_task(
                lead["id"],
                f"Reply received — respond to {lead['first_name']} within 1 hour",
                (datetime.utcnow() + timedelta(hours=1)).isoformat(),
                priority="High",
            )
            added += 1

    flash(f"Sync complete: {added} new reply(ies) found.", "success" if added > 0 else "info")
    return redirect(url_for("replies"))


@app.route("/leads/<int:lead_id>/log-reply", methods=["POST"])
def log_reply(lead_id):
    lead     = db.get_lead(lead_id)
    subject  = request.form.get("subject", "Reply from lead")
    body     = request.form.get("body", "")
    priority = request.form.get("priority", "Medium")

    db.add_reply(lead_id, lead["email"], subject, body, priority, source="Manual")
    db.update_lead_status(lead_id, "Engaged")
    db.create_task(
        lead_id,
        f"Reply received — respond to {lead['first_name']} within 1 hour",
        (datetime.utcnow() + timedelta(hours=1)).isoformat(),
        priority="High",
    )
    flash(f"Reply logged as {priority} priority!", "success")
    return redirect(url_for("lead_detail", lead_id=lead_id))


@app.route("/replies/<int:reply_id>/priority", methods=["POST"])
def update_reply_priority(reply_id):
    priority = request.form.get("priority")
    db.update_reply(reply_id, priority=priority)
    return redirect(url_for("replies"))


@app.route("/replies/<int:reply_id>/status", methods=["POST"])
def update_reply_status(reply_id):
    status = request.form.get("status")
    db.update_reply(reply_id, status=status)
    return redirect(request.referrer or url_for("replies"))


@app.route("/replies/<int:reply_id>/delete", methods=["POST"])
def delete_reply(reply_id):
    db.delete_reply(reply_id)
    flash("Reply deleted.", "info")
    return redirect(url_for("replies"))


# ── IMPORT FROM EXCEL ──────────────────────────────────────────────────────────

@app.route("/leads/import", methods=["GET", "POST"])
def import_leads():
    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file:
            flash("No file selected.", "danger")
            return redirect(url_for("import_leads"))
        try:
            from openpyxl import load_workbook
            import io

            wb   = load_workbook(io.BytesIO(file.read()))
            # Try sheet named "Lead Database", fall back to first sheet
            ws   = wb["Lead Database"] if "Lead Database" in wb.sheetnames else wb.active

            # Read header row (row 2 = index 1, since header=1 in pandas)
            rows     = list(ws.iter_rows(values_only=True))
            # Find header row — first row that contains "Email"
            hdr_idx  = 0
            for i, r in enumerate(rows):
                if r and any(str(c).strip().lower() == "email" for c in r if c):
                    hdr_idx = i
                    break
            headers  = [str(c).strip() if c else "" for c in rows[hdr_idx]]

            def col(row_vals, name):
                try:
                    idx = headers.index(name)
                    v   = row_vals[idx]
                    return str(v).strip() if v is not None else ""
                except (ValueError, IndexError):
                    return ""

            imported, skipped = 0, 0

            for row_vals in rows[hdr_idx + 1:]:
                if not any(row_vals):
                    continue
                try:
                    name  = col(row_vals, "Decision Maker") or "Unknown"
                    parts = name.split()
                    tmpl  = col(row_vals, "Email Template").split("—")[0].strip().replace("Template ", "")
                    data  = {
                        "first_name"            : parts[0] if parts else "Unknown",
                        "last_name"             : " ".join(parts[1:]) if len(parts) > 1 else "",
                        "email"                 : col(row_vals, "Email"),
                        "company"               : col(row_vals, "Company Name"),
                        "title"                 : col(row_vals, "Title"),
                        "phone"                 : "",
                        "website"               : col(row_vals, "Website"),
                        "city"                  : col(row_vals, "City / State").split(",")[0].strip(),
                        "country"               : col(row_vals, "Country") or "USA",
                        "industry"              : col(row_vals, "Company Type"),
                        "status"                : "New",
                        "priority_score"        : int(col(row_vals, "Priority Score") or 0),
                        "services_needed"       : col(row_vals, "Services Needed"),
                        "outsourcing_likelihood": col(row_vals, "Outsourcing Likelihood"),
                        "pitch_angle"           : col(row_vals, "Pitch Angle"),
                        "email_template"        : tmpl if tmpl in ("A", "B", "C", "D") else "A",
                        "linkedin_url"          : col(row_vals, "LinkedIn URL"),
                        "follow_up_stage"       : "",
                        "description"           : col(row_vals, "Pain Point"),
                    }
                    if not data["email"] or data["email"].lower() in ("nan", ""):
                        skipped += 1
                        continue
                    db.create_lead(data)
                    imported += 1
                except Exception:
                    skipped += 1

            flash(f"Import complete: {imported} imported, {skipped} skipped.", "success")
        except Exception as e:
            flash(f"Import failed: {str(e)}", "danger")
        return redirect(url_for("leads"))

    return render_template("import.html")


# ── SETTINGS ───────────────────────────────────────────────────────────────────

@app.route("/settings")
def settings():
    cid  = os.getenv("ZOHO_CLIENT_ID", "NOT SET")
    dc   = os.getenv("ZOHO_DC", "in")
    acct = os.getenv("ZOHO_MAIL_ACCOUNT_ID", "NOT SET")
    hkey = os.getenv("HUNTER_API_KEY", "")
    masked_cid  = cid[:10] + "..." if len(cid) > 10 else cid
    masked_hkey = hkey[:8] + "..." if len(hkey) > 8 else ("" if not hkey else hkey)
    return render_template("settings.html",
                           zoho_client_id=masked_cid,
                           zoho_dc=dc,
                           zoho_account_id=acct,
                           hunter_key=masked_hkey)


TEMPLATE_LABELS = {
    "A": {"name": "Template A", "desc": "Enterprise GC / Engineering Firms (USA/Canada/AU)", "icon": "bi-building"},
    "B": {"name": "Template B", "desc": "Architecture Firms (USA/Germany/Europe)",            "icon": "bi-pencil-ruler"},
    "C": {"name": "Template C", "desc": "Drone / Scan-to-BIM / Survey Firms",                 "icon": "bi-camera"},
    "D": {"name": "Template D", "desc": "INFRA X / Geospatial / Survey (India)",               "icon": "bi-geo-alt"},
}


@app.route("/templates", methods=["GET"])
@login_required
def templates_list():
    active = request.args.get("t", "A")
    if active not in TEMPLATE_LABELS:
        active = "A"

    # Load saved DB templates into lookup: {(key,step): row}
    saved = db.get_all_email_templates()

    # Build display data: for active template, get subject/body for each step (DB or default)
    steps_display = []
    for step in range(3):
        db_row = saved.get((active, step))
        if db_row:
            subject   = db_row["subject"]
            body      = db_row["body"]
            is_custom = True
        else:
            body      = _get_raw_body(active, step)
            # Extract subject from built-in default via raw body lookup
            subject   = _get_default_subject(active, step)
            is_custom = False
        steps_display.append({"step": step, "subject": subject, "body": body, "is_custom": is_custom})

    return render_template("templates.html",
                           templates=TEMPLATE_LABELS,
                           active=active,
                           steps=steps_display)


@app.route("/templates/save", methods=["POST"])
@login_required
def save_template():
    key     = request.form.get("template_key", "A")
    step    = int(request.form.get("step", 0))
    subject = request.form.get("subject", "").strip()
    body    = request.form.get("body", "").strip()
    if key not in TEMPLATE_LABELS or step not in (0, 1, 2):
        flash("Invalid template.", "danger")
        return redirect(url_for("templates_list"))
    db.save_email_template(key, step, subject, body)
    flash(f"Template {key} — Step {step+1} saved!", "success")
    return redirect(url_for("templates_list", t=key))


@app.route("/templates/reset", methods=["POST"])
@login_required
def reset_template():
    key  = request.form.get("template_key", "A")
    step = int(request.form.get("step", 0))
    conn = db.get_db()
    c = conn.cursor()
    c.execute(db._q("DELETE FROM email_templates WHERE template_key=? AND step=?"), (key, step))
    conn.commit()
    conn.close()
    flash(f"Template {key} — Step {step+1} reset to default.", "info")
    return redirect(url_for("templates_list", t=key))


def _get_default_subject(template_key: str, step: int) -> str:
    subjects = {
        "A": [
            "{first_name}, is {company}'s BIM team at full capacity right now?",
            "Re: BIM capacity at {company} — one example that might be relevant",
            "Last note — pilot slots for {company}",
        ],
        "B": [
            "{first_name} — is your Revit team keeping up with project deadlines at {company}?",
            "{first_name} — our work for firms like {company} (quick examples)",
            "Final note from me — {company}",
        ],
        "C": [
            "{first_name} — turning {company}'s scan data into BIM models (we've done 15+ projects)",
            "{first_name} — our Scan-to-BIM workflow for {company} (short overview)",
            "Last note — Scan-to-BIM pilot for {company}",
        ],
        "D": [
            "{first_name}, one question about how {company} monitors site progress",
            "INFRA X at Khavada — what {company} can do with this",
            "Last note — INFRA X and {company}",
        ],
    }
    return subjects.get(template_key, subjects["A"])[step]


def _get_raw_body(template_key: str, step: int) -> str:
    """Extract raw (unwrapped) body from built-in zoho_mail templates."""
    from zoho_mail import (PORTFOLIO_LINK, DRONE_PORTFOLIO_LINK, CALENDLY_LINK,
                            WEBSITE, PHONE, TITLE, COMPANY_FULL)
    fn, co = "{first_name}", "{company}"
    pl = DRONE_PORTFOLIO_LINK if template_key in ("C", "D") else PORTFOLIO_LINK

    A = [
        f"""<p>Hi {{first_name}},</p>
<p>I came across {{company}} and noticed you're handling significant construction/engineering projects — I wanted to reach out directly.</p>
<p>I'm Kishan, Co-founder &amp; CFO at <strong>BIM INFRASOLUTIONS LLP</strong>. We're a technology-driven BIM consultancy with <strong>100+ projects delivered across USA, Canada, Germany, Australia and India</strong>.</p>
<p>A quick question: when your BIM team hits capacity during peak project phases, what does {{company}} typically do — hire temps, delay, or push the in-house team harder?</p>
<p>We step in as a seamless offshore extension:</p>
<ul style="line-height:2;">
    <li>Revit modelling — Architectural, Structural, MEP (LOD 200–500)</li>
    <li>Clash detection &amp; BIM coordination (BIM 360 / ACC / Procore)</li>
    <li>4D scheduling &amp; 5D quantity takeoffs</li>
    <li>Scan-to-BIM from LiDAR / point clouds</li>
    <li>Dynamo / automation scripts</li>
</ul>
<p>We're offering <strong>3 paid pilot slots this quarter</strong> — results in 48 hours, no long-term commitment.</p>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Book a 30-Min Call</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>Following up — wanted to share a project directly relevant to {{company}}.</p>
<p>We recently completed BIM coordination for a <strong>high-rise in Canada</strong> (6,000 sqmt, LOD 300, full ARCHI + MEPF + Structure) — delivered in under 2 weeks, clash-free.</p>
<p>We've also done a <strong>110 km floating tunnel for Parsons</strong> with a Dynamo-based automated alignment system that saved weeks of manual rework.</p>
<p><strong>3 things that set us apart:</strong></p>
<ul style="line-height:2;">
    <li>360,000+ hours of BIM work delivered</li>
    <li>54 million sqmt of assets digitalized</li>
    <li>Onboarding in 48 hours — we match your Revit templates &amp; standards</li>
</ul>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Schedule 30 Minutes</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>I'll keep this short — this is my last note.</p>
<p>If BIM outsourcing ever comes up at {{company}} — whether for a single project, a capacity crunch, or an ongoing engagement — I'd love to be your first call.</p>
<p>We have <strong>2 pilot slots remaining this quarter</strong> — a real project deliverable at a fixed rate, zero risk.</p>
<p><a href="{CALENDLY_LINK}" style="background:#D4A017;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Book a Final 30-Min Call</a></p>
<p>Wishing {{company}} a great quarter ahead.</p>""",
    ]
    B = [
        f"""<p>Hi {{first_name}},</p>
<p>Architecture firms like {{company}} often deal with Revit workload peaks during design development and CD phases — the in-house team gets stretched thin.</p>
<p>I'm Kishan, Co-founder &amp; CFO at <strong>BIM INFRASOLUTIONS LLP</strong>. We work as a dedicated BIM extension for architecture firms across USA, Germany, and Europe.</p>
<p>Recent architecture work:</p>
<ul style="line-height:2;">
    <li><strong>Reutlingen City Hall, Germany</strong> — 70,000 sqmt full BIM model</li>
    <li><strong>Nikko Hotel Düsseldorf</strong> — 40,000 sqmt Scan-to-BIM + coordination</li>
    <li><strong>Bonn University</strong> — 30,000 sqmt LOD 400 Scan-to-BIM</li>
    <li>60+ parametric Revit families (LOD 500)</li>
</ul>
<p>We work inside your standards — your Revit templates, your sheet naming, your family library.</p>
<p><strong>Deliverable in 48 hours. Pilot project available this quarter.</strong></p>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Book a 30-Min Call</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>Following up — quick specifics on how we've helped firms similar to {{company}}.</p>
<p><strong>What we handle:</strong></p>
<ul style="line-height:2;">
    <li>Revit CD documentation — floor plans, sections, elevations, details</li>
    <li>Revit family creation (parametric, LOD 300–500)</li>
    <li>BIM coordination — clash detection across ARCHI, MEPF, Structure</li>
    <li>As-built to BIM (Scan-to-BIM, LOD 300–400)</li>
    <li>BIM 360 / ACC project setup and management</li>
</ul>
<p>100+ international projects | 54M sqmt digitalized | 360K hours delivered</p>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Let's Talk — 30 Min Call</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>Last note — I won't keep following up after this.</p>
<p>If {{company}} ever needs BIM/Revit support — CD crunch, Scan-to-BIM, or coordination capacity — we're ready within 48 hours.</p>
<p>Pilot offer: <strong>one real project deliverable at a fixed rate</strong>, only continue if satisfied.</p>
<p><a href="{CALENDLY_LINK}" style="background:#D4A017;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Book a Call Anytime</a></p>
<p>Best of luck with your current projects, {{first_name}}.</p>""",
    ]
    C = [
        f"""<p>Hi {{first_name}},</p>
<p>If {{company}} works with drone surveys, LiDAR scans, or point cloud data — we convert raw scan data into accurate, fully coordinated BIM models.</p>
<p>Recent Scan-to-BIM projects:</p>
<ul style="line-height:2;">
    <li><strong>Rochsburg Castle, Germany</strong> — 40,000 sqmt heritage, LOD 300</li>
    <li><strong>Kampnagel Hamburg</strong> — 100,000 sqmt industrial renovation</li>
    <li><strong>Bonn University</strong> — 30,000 sqmt, LOD 400</li>
    <li><strong>Autobahn Viaduct</strong> — 500m historic bridge</li>
    <li><strong>Nikko Hotel Düsseldorf</strong> — 40,000 sqmt full renovation</li>
</ul>
<p>Delivery: raw point cloud to finished Revit model in <strong>3–5 business days</strong>.</p>
<p><strong>Paid pilot:</strong> send a sample scan, we deliver a BIM model. Pay only if satisfied.</p>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Book a 30-Min Demo Call</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>Quick follow-up — our standard Scan-to-BIM workflow for {{company}}'s projects:</p>
<ol style="line-height:2;">
    <li><strong>Receive</strong> — point cloud / E57 / RCP / drone data</li>
    <li><strong>Register &amp; clean</strong> — align scans, remove noise</li>
    <li><strong>Model</strong> — Revit model to your LOD (200–500)</li>
    <li><strong>QC &amp; deliver</strong> — clash-free, your naming convention</li>
</ol>
<p>Turnaround: <strong>3–5 business days</strong>. We've digitalized <strong>54 million sqmt</strong> globally.</p>
<p>Happy to do a <strong>free test conversion</strong> on a small sample scan.</p>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">Schedule 30 Minutes</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>This is my last follow-up. If {{company}} has any scan data needing BIM conversion, or upcoming survey projects needing BIM deliverables, we'd love to be your partner.</p>
<p><strong>Paid pilot, fixed price, satisfaction guaranteed. No lock-in.</strong></p>
<p><a href="{CALENDLY_LINK}" style="background:#D4A017;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:8px;">One Last Call — 30 Min</a></p>
<p>Wishing {{company}} great projects ahead.</p>""",
    ]
    D = [
        f"""<p>Hi {{first_name}},</p>
<p>{{company}}'s geospatial/survey work aligns closely with <strong>INFRA X</strong> — our site intelligence platform for large infrastructure projects.</p>
<p><strong>Key capabilities:</strong></p>
<ul style="line-height:2;">
    <li><strong>Compare View</strong> — planned design vs. actual site condition, week by week</li>
    <li><strong>One-desk monitoring</strong> — all sites, all zones, all teams on one dashboard</li>
    <li><strong>Resource tracking</strong> — manpower, machinery, materials automatically logged</li>
    <li><strong>Automated reports</strong> — volume calculations, milestone %, deviation alerts</li>
    <li><strong>As-built documentation</strong> — orthomosaic maps, point clouds, 3D models</li>
</ul>
<p>Currently <strong>live at Khavada</strong> for <strong>Kalpataru Projects</strong> and <strong>Hitachi Energy India</strong>.</p>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;">Book a 30-Min Call</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>Following up on INFRA X — here's what's running live at <strong>Khavada</strong> for Kalpataru and Hitachi Energy:</p>
<ul style="line-height:2;">
    <li>Compare View — any date, planned vs actual</li>
    <li>One-desk monitoring — 500+ acres, one screen, updated weekly</li>
    <li>Resource allocation — equipment, labour zones, materials tracked</li>
    <li>Deviation alerts — flags issues before weekly review</li>
    <li>Auto-generated client reports — every Monday morning</li>
</ul>
<p>For {{company}}, INFRA X adds the intelligence layer on top of your geospatial capability.</p>
<p>We have <strong>3 partnership slots open</strong> with Survey of India empanelled firms this quarter.</p>
<p><a href="{CALENDLY_LINK}" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;">Schedule a 30-Min Demo</a></p>""",
        f"""<p>Hi {{first_name}},</p>
<p>Last note — brief.</p>
<p>If {{company}} works on large infrastructure or renewable energy sites where clients struggle with ground progress visibility — <strong>INFRA X solves exactly that</strong>.</p>
<p>One dashboard. Compare view. Resource allocation. Automated reports. Live at Khavada with Kalpataru and Hitachi Energy — a working, deployed solution.</p>
<p>Open to partnership, white-label, or referral — whatever works for {{company}}.</p>
<p><a href="{CALENDLY_LINK}" style="background:#D4A017;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;">One Final Call — 30 Min</a></p>
<p>Wishing {{company}} great projects ahead.</p>""",
    ]
    mapping = {"A": A, "B": B, "C": C, "D": D}
    steps = mapping.get(template_key, A)
    return steps[step] if step < len(steps) else ""


@app.route("/my-settings", methods=["GET", "POST"])
@login_required
def my_settings():
    username = current_user.username
    cfg = db.get_user_settings(username) or {}

    if request.method == "POST":
        action = request.form.get("action", "save")

        # If settings are locked, only allow unlock
        if cfg.get("is_locked") and action != "unlock":
            flash("Settings are locked. Click Unlock to edit.", "warning")
            return redirect(url_for("my_settings"))

        if action == "unlock":
            cfg["is_locked"] = 0
            db.save_user_settings(username, {
                "sender_email"      : cfg.get("sender_email", ""),
                "sender_name"       : cfg.get("sender_name", ""),
                "zoho_client_id"    : cfg.get("zoho_client_id", ""),
                "zoho_client_secret": cfg.get("zoho_client_secret", ""),
                "zoho_refresh_token": cfg.get("zoho_refresh_token", ""),
                "zoho_dc"           : cfg.get("zoho_dc", "in"),
                "zoho_account_id"   : cfg.get("zoho_account_id", ""),
                "is_locked"         : 0,
                "wa_phone"          : cfg.get("wa_phone", ""),
                "callmebot_api_key" : cfg.get("callmebot_api_key", ""),
            })
            flash("Settings unlocked. You can now edit.", "info")
            return redirect(url_for("my_settings"))

        # Save
        lock = 1 if request.form.get("lock") == "on" else 0
        data = {
            "sender_email"      : request.form.get("sender_email", "").strip(),
            "sender_name"       : request.form.get("sender_name", "").strip(),
            "zoho_client_id"    : request.form.get("zoho_client_id", "").strip(),
            "zoho_client_secret": request.form.get("zoho_client_secret", "").strip(),
            "zoho_refresh_token": request.form.get("zoho_refresh_token", "").strip(),
            "zoho_dc"           : request.form.get("zoho_dc", "in").strip(),
            "zoho_account_id"   : request.form.get("zoho_account_id", "").strip(),
            "is_locked"         : lock,
            "wa_phone"          : request.form.get("wa_phone", "").strip(),
            "callmebot_api_key" : request.form.get("callmebot_api_key", "").strip(),
        }
        db.save_user_settings(username, data)
        flash("Settings saved and locked!" if lock else "Settings saved.", "success")
        return redirect(url_for("my_settings"))

    return render_template("my_settings.html", cfg=cfg)


@app.route("/test-send-email")
@login_required
def test_send_email():
    """Send a test email to the logged-in user's own address to verify Zoho credentials."""
    username = current_user.username
    user_cfg = db.get_user_settings(username) or {}
    to_email = user_cfg.get("sender_email") or TEAM_EMAILS.get(username)
    if not to_email:
        flash("No sender email set. Configure your email in My Email Settings first.", "danger")
        return redirect(url_for("my_settings"))
    try:
        ok = mail_client.send_email(
            to_address=to_email,
            subject="BIM CRM — Test Email ✅",
            html_body=f"""
            <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:480px;margin:20px auto;padding:24px;
                        border:1px solid #e0e0e0;border-radius:8px;color:#222;">
              <div style="background:#1B3A6B;color:#fff;padding:14px 20px;border-radius:6px 6px 0 0;margin:-24px -24px 20px -24px;">
                <strong>BIM Infra Solutions CRM — Test Email</strong>
              </div>
              <p>Hi <strong>{current_user.display}</strong>,</p>
              <p>Your Zoho Mail credentials are working correctly.
                 This test was sent from the BIM CRM using your personal credentials.</p>
              <p style="color:#888;font-size:13px;">Sender: {to_email}</p>
            </div>""",
            user_settings=user_cfg or None,
        )
        if ok:
            flash(f"Test email sent to {to_email} — check your inbox!", "success")
        else:
            flash(f"Zoho returned failure. Check your credentials in My Email Settings (token may be expired).", "danger")
    except Exception as e:
        err = str(e)
        if "invalid_code" in err or "invalid_token" in err or "refresh" in err.lower():
            flash("Zoho token expired — run get_token.py to get a new refresh token.", "danger")
        else:
            flash(f"Test email error: {err}", "danger")
    return redirect(url_for("my_settings"))


@app.route("/test-whatsapp-bot")
@login_required
def test_whatsapp_bot():
    """Send a test WhatsApp message via CallMeBot to verify setup."""
    username = current_user.username
    _send_whatsapp_bot(username,
        f"✅ BIM CRM WhatsApp Bot is working!\nHi {current_user.display}, your notifications are active.")
    flash("Test WhatsApp sent! Check your phone. If not received, verify your phone number and API key.", "info")
    return redirect(url_for("my_settings"))


@app.route("/test-connection")
def test_connection():
    result = mail_client.test_connection()
    if result["ok"]:
        flash(f"Zoho Mail connected! Found {result['accounts']} account(s). Token: {result['token']}", "success")
    else:
        flash(f"Connection failed: {result['error']}", "danger")
    return redirect(url_for("settings"))


# ── API ────────────────────────────────────────────────────────────────────────

@app.route("/api/stats")
def api_stats():
    return jsonify(db.get_stats())


@app.route("/api/hunter/find/<int:lead_id>")
def hunter_find(lead_id):
    """Find email for a lead via Hunter.io Email Finder API."""
    import requests as req
    api_key = os.getenv("HUNTER_API_KEY", "")
    if not api_key:
        return jsonify({"error": "Hunter.io API key not set in .env"}), 400

    lead = db.get_lead(lead_id)
    if not lead:
        return jsonify({"error": "Lead not found"}), 404

    # Extract domain from website
    website = lead.get("website", "")
    domain  = website.replace("https://","").replace("http://","").replace("www.","").strip("/").split("/")[0]
    if not domain:
        return jsonify({"error": "No website/domain set for this lead"}), 400

    first = lead.get("first_name", "")
    last  = lead.get("last_name", "")

    r = req.get("https://api.hunter.io/v2/email-finder", params={
        "domain"    : domain,
        "first_name": first,
        "last_name" : last,
        "api_key"   : api_key,
    }, timeout=10)

    data = r.json()
    if r.status_code != 200 or "data" not in data:
        return jsonify({"error": data.get("errors", [{"details": "Hunter API error"}])[0].get("details", "Unknown error")}), 400

    result = data["data"]
    return jsonify({
        "email"      : result.get("email"),
        "score"      : result.get("score"),
        "sources"    : len(result.get("sources", [])),
        "first_name" : first,
        "last_name"  : last,
        "domain"     : domain,
    })


@app.route("/api/hunter/verify")
def hunter_verify():
    """Verify an email address via Hunter.io."""
    import requests as req
    api_key = os.getenv("HUNTER_API_KEY", "")
    if not api_key:
        return jsonify({"error": "Hunter.io API key not set in .env"}), 400

    email = request.args.get("email", "")
    if not email:
        return jsonify({"error": "No email provided"}), 400

    r = req.get("https://api.hunter.io/v2/email-verifier", params={
        "email"  : email,
        "api_key": api_key,
    }, timeout=15)

    data = r.json()
    if r.status_code != 200 or "data" not in data:
        return jsonify({"error": "Verification failed"}), 400

    result = data["data"]
    return jsonify({
        "email"    : result.get("email"),
        "status"   : result.get("status"),       # valid / invalid / accept_all / unknown
        "score"    : result.get("score"),
        "regexp"   : result.get("regexp"),
        "mx_records": result.get("mx_records"),
        "smtp_server": result.get("smtp_server"),
        "smtp_check": result.get("smtp_check"),
    })


@app.route("/api/hunter/domain/<int:lead_id>")
def hunter_domain(lead_id):
    """Search all emails for a company domain via Hunter.io."""
    import requests as req
    api_key = os.getenv("HUNTER_API_KEY", "")
    if not api_key:
        return jsonify({"error": "Hunter.io API key not set in .env"}), 400

    lead   = db.get_lead(lead_id)
    website = lead.get("website", "") if lead else ""
    domain  = website.replace("https://","").replace("http://","").replace("www.","").strip("/").split("/")[0]
    if not domain:
        return jsonify({"error": "No website/domain set for this lead"}), 400

    r = req.get("https://api.hunter.io/v2/domain-search", params={
        "domain" : domain,
        "api_key": api_key,
        "limit"  : 10,
    }, timeout=10)

    data = r.json()
    if r.status_code != 200 or "data" not in data:
        return jsonify({"error": "Domain search failed"}), 400

    result  = data["data"]
    emails  = result.get("emails", [])
    pattern = result.get("pattern", "")
    return jsonify({
        "domain"      : domain,
        "pattern"     : pattern,
        "total"       : result.get("total", 0),
        "emails"      : [{"value": e.get("value"), "type": e.get("type"), "confidence": e.get("confidence"), "first_name": e.get("first_name"), "last_name": e.get("last_name"), "position": e.get("position")} for e in emails[:10]],
    })


@app.route("/leads/<int:lead_id>/update-email", methods=["POST"])
def update_email(lead_id):
    new_email = request.form.get("email", "").strip()
    if new_email:
        try:
            db.update_lead(lead_id, {"email": new_email})
            # Reset sequence step so email can be resent
            db.update_lead(lead_id, {"email_sequence_step": 0})
            db.update_lead_status(lead_id, "New")
            flash(f"Email updated to {new_email}. Sequence reset — ready to send.", "success")
        except Exception as e:
            flash(f"Could not update email: {e}", "danger")
    return redirect(url_for("lead_detail", lead_id=lead_id))


@app.route("/api/import-leads", methods=["POST"])
def api_import_leads():
    """Bulk import leads via JSON POST — runs inside Flask process to avoid DB lock."""
    import json as _json
    data = request.get_json(force=True)
    leads_data = data.get("leads", [])
    imported, skipped = 0, 0
    for lead in leads_data:
        lead.setdefault("phone", "")
        lead.setdefault("follow_up_stage", "")
        lead.setdefault("status", "New")
        try:
            db.create_lead(lead)
            imported += 1
        except Exception:
            skipped += 1
    return jsonify({"imported": imported, "skipped": skipped, "total": imported + skipped})


# ── HELPERS ────────────────────────────────────────────────────────────────────

def _form_to_lead(form) -> dict:
    return {
        "first_name"            : form.get("first_name", ""),
        "last_name"             : form.get("last_name", ""),
        "email"                 : form.get("email", ""),
        "company"               : form.get("company", ""),
        "title"                 : form.get("title", ""),
        "phone"                 : form.get("phone", ""),
        "website"               : form.get("website", ""),
        "city"                  : form.get("city", ""),
        "country"               : form.get("country", "USA"),
        "industry"              : form.get("industry", ""),
        "priority_score"        : int(form.get("priority_score", 0) or 0),
        "services_needed"       : form.get("services_needed", ""),
        "outsourcing_likelihood": form.get("outsourcing_likelihood", ""),
        "pitch_angle"           : form.get("pitch_angle", ""),
        "email_template"        : form.get("email_template", "A"),
        "linkedin_url"          : form.get("linkedin_url", ""),
        "follow_up_stage"       : "",
        "description"           : form.get("description", ""),
    }


# ── TEAM EMAIL MAP ────────────────────────────────────────────────────────────

TEAM_EMAILS = {
    "kishan"  : "kishan.batavia@biminfrasolutions.in",
    "hirakraj": "ceo@biminfrasolutions.com",
    "tirth"   : "coo@biminfrasolutions.com",
    "jenish"  : "services@biminfrasolutions.in",
}

TEAM_DISPLAY = {k: v["display"] for k, v in CRM_USERS.items()}


def _send_internal_email(to_username: str, subject: str, html_body: str):
    """Send an internal notification email to a team member."""
    to_email = TEAM_EMAILS.get(to_username)
    if not to_email:
        return
    try:
        mail_client.send_email(to_address=to_email, subject=subject, html_body=html_body)
    except Exception as e:
        app.logger.warning("Internal email to %s failed: %s", to_username, e)


def _send_whatsapp_bot(username: str, message: str):
    """Send an automated WhatsApp message via CallMeBot API to a team member."""
    import urllib.parse, requests as req
    cfg = db.get_user_settings(username)
    if not cfg:
        return
    phone  = cfg.get("wa_phone", "")
    apikey = cfg.get("callmebot_api_key", "")
    if not phone or not apikey:
        return
    try:
        encoded = urllib.parse.quote(message)
        url = f"https://api.callmebot.com/whatsapp.php?phone={phone}&text={encoded}&apikey={apikey}"
        req.get(url, timeout=10)
    except Exception as e:
        app.logger.warning("CallMeBot WA to %s failed: %s", username, e)


def _task_email_html(task: dict, action: str = "assigned") -> str:
    assigned_by_display = TEAM_DISPLAY.get(task.get("assigned_by", ""), task.get("assigned_by", ""))
    due = str(task.get("due_date", ""))[:10] if task.get("due_date") else "No due date"
    return f"""
    <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:520px;margin:20px auto;padding:0 20px;color:#222;font-size:14px;line-height:1.8;">
      <div style="background:#1B3A6B;color:#fff;padding:16px 22px;border-radius:8px 8px 0 0;">
        <strong>BIM Infra Solutions — Internal CRM</strong>
      </div>
      <div style="border:1px solid #e0e0e0;border-top:none;padding:22px;border-radius:0 0 8px 8px;">
        <h3 style="color:#1B3A6B;margin-top:0;">Task {action.title()}</h3>
        <p><strong>{task['title']}</strong></p>
        {f"<p style='color:#555;'>{task.get('description','')}</p>" if task.get('description') else ''}
        <table style="font-size:13px;color:#555;border-collapse:collapse;width:100%;">
          <tr><td style="padding:4px 0;width:120px;"><strong>Assigned by:</strong></td><td>{assigned_by_display}</td></tr>
          <tr><td style="padding:4px 0;"><strong>Priority:</strong></td><td>{task.get('priority','Medium')}</td></tr>
          <tr><td style="padding:4px 0;"><strong>Due date:</strong></td><td>{due}</td></tr>
        </table>
        <p style="margin-top:18px;">
          <a href="https://app.biminfrasolutions.in/team-tasks"
             style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;">
            View Task in CRM
          </a>
        </p>
      </div>
    </div>"""


# ── INTERNAL NOTES (Instructions) ─────────────────────────────────────────────

@app.route("/instructions", methods=["GET", "POST"])
@login_required
def instructions():
    if request.method == "POST":
        action = request.form.get("action", "create")
        if action == "create":
            db.create_note(
                title    = request.form.get("title","").strip(),
                body     = request.form.get("body","").strip(),
                created_by = current_user.username,
                is_pinned  = 1 if request.form.get("pin") else 0,
            )
            flash("Instruction posted!", "success")
        elif action == "edit":
            db.update_note(
                int(request.form.get("note_id")),
                request.form.get("title","").strip(),
                request.form.get("body","").strip(),
                1 if request.form.get("pin") else 0,
            )
            flash("Updated.", "success")
        elif action == "delete":
            db.delete_note(int(request.form.get("note_id")))
            flash("Deleted.", "info")
        elif action == "toggle_pin":
            note_id = int(request.form.get("note_id"))
            notes = db.get_notes()
            note  = next((n for n in notes if n["id"] == note_id), None)
            if note:
                db.update_note(note_id, note["title"], note["body"], 0 if note["is_pinned"] else 1)
        return redirect(url_for("instructions"))

    notes = db.get_notes()
    return render_template("instructions.html", notes=notes, team=TEAM_DISPLAY)


# ── RESPONSIBILITIES ───────────────────────────────────────────────────────────

@app.route("/responsibilities", methods=["GET", "POST"])
@login_required
def responsibilities():
    filter_user = request.args.get("user", "")
    if request.method == "POST":
        action = request.form.get("action", "create")
        if action == "create":
            assigned_to = request.form.get("assigned_to")
            data = {
                "title"      : request.form.get("title","").strip(),
                "description": request.form.get("description","").strip(),
                "assigned_to": assigned_to,
                "assigned_by": current_user.username,
                "category"   : request.form.get("category","General"),
                "status"     : "Active",
            }
            db.create_responsibility(data)
            # Notify the assigned person
            display_to   = TEAM_DISPLAY.get(assigned_to, assigned_to)
            display_from = TEAM_DISPLAY.get(current_user.username, current_user.username)
            _send_internal_email(
                assigned_to,
                f"New Responsibility Assigned — {data['title']}",
                f"""<div style="font-family:'Segoe UI',Arial,sans-serif;max-width:520px;margin:20px auto;padding:0 20px;font-size:14px;line-height:1.8;color:#222;">
                  <div style="background:#1B3A6B;color:#fff;padding:16px 22px;border-radius:8px 8px 0 0;"><strong>BIM Infra Solutions CRM</strong></div>
                  <div style="border:1px solid #e0e0e0;border-top:none;padding:22px;border-radius:0 0 8px 8px;">
                    <h3 style="color:#1B3A6B;margin-top:0;">Responsibility Assigned to You</h3>
                    <p><strong>{data['title']}</strong></p>
                    {f"<p style='color:#555;'>{data['description']}</p>" if data['description'] else ''}
                    <p style='color:#555;'>Category: <strong>{data['category']}</strong><br>Assigned by: <strong>{display_from}</strong></p>
                    <a href="https://app.biminfrasolutions.in/responsibilities" style="background:#1B3A6B;color:#fff;padding:10px 20px;border-radius:5px;text-decoration:none;display:inline-block;margin-top:10px;">View in CRM</a>
                  </div></div>"""
            )
            # WhatsApp bot notification
            wa_msg = f"📋 New Responsibility: *{data['title']}*\nCategory: {data['category']}\nAssigned by: {display_from}\nhttps://app.biminfrasolutions.in/responsibilities"
            if assigned_to == "all":
                for uname in TEAM_EMAILS:
                    _send_whatsapp_bot(uname, wa_msg)
            else:
                _send_whatsapp_bot(assigned_to, wa_msg)
            flash(f"Responsibility assigned to {display_to} and notified by email + WhatsApp.", "success")
        elif action == "status":
            db.update_responsibility_status(int(request.form.get("rid")), request.form.get("status"))
            flash("Status updated.", "success")
        elif action == "delete":
            db.delete_responsibility(int(request.form.get("rid")))
            flash("Deleted.", "info")
        return redirect(url_for("responsibilities", user=filter_user))

    all_resp  = db.get_responsibilities(assigned_to=filter_user or None)
    # Group by user for display
    grouped = {}
    for r in all_resp:
        u = r["assigned_to"]
        grouped.setdefault(u, []).append(r)
    return render_template("responsibilities.html",
                           grouped=grouped, all_resp=all_resp,
                           team=TEAM_DISPLAY, filter_user=filter_user)


# ── TEAM TASKS ─────────────────────────────────────────────────────────────────

@app.route("/team-tasks", methods=["GET", "POST"])
@login_required
def team_tasks():
    filter_user   = request.args.get("user", "")
    filter_status = request.args.get("status", "")

    if request.method == "POST":
        action = request.form.get("action", "create")
        if action == "create":
            assigned_to = request.form.get("assigned_to","")
            lead_id_raw = request.form.get("lead_id","").strip()
            due_raw     = request.form.get("due_date","").strip()
            rem_raw     = request.form.get("reminder_at","").strip()
            data = {
                "title"      : request.form.get("title","").strip(),
                "description": request.form.get("description","").strip(),
                "assigned_to": assigned_to,
                "assigned_by": current_user.username,
                "lead_id"    : int(lead_id_raw) if lead_id_raw.isdigit() else None,
                "due_date"   : due_raw or None,
                "priority"   : request.form.get("priority","Medium"),
                "status"     : "Pending",
                "reminder_at": rem_raw or None,
            }
            db.create_team_task(data)
            # Send email notification to all if 'all', else specific user
            targets = list(TEAM_EMAILS.keys()) if assigned_to == "all" else [assigned_to]
            for username in targets:
                if username != current_user.username:
                    _send_internal_email(
                        username,
                        f"Task Assigned to You — {data['title']}",
                        _task_email_html(data, "assigned"),
                    )
            display_to = "Everyone" if assigned_to == "all" else TEAM_DISPLAY.get(assigned_to, assigned_to)
            # WhatsApp bot notification
            wa_msg = (f"✅ Task Assigned: *{data['title']}*\n"
                      f"Priority: {data['priority']} | Due: {data['due_date'] or 'No date'}\n"
                      f"By: {TEAM_DISPLAY.get(current_user.username, current_user.username)}\n"
                      f"https://app.biminfrasolutions.in/team-tasks")
            wa_targets = list(TEAM_EMAILS.keys()) if assigned_to == "all" else [assigned_to]
            for uname in wa_targets:
                if uname != current_user.username:
                    _send_whatsapp_bot(uname, wa_msg)
            flash(f"Task assigned to {display_to} and notified by email + WhatsApp.", "success")

        elif action == "status":
            db.update_team_task_status(int(request.form.get("tid")), request.form.get("status"))
            flash("Task status updated.", "success")

        elif action == "delete":
            db.delete_team_task(int(request.form.get("tid")))
            flash("Task deleted.", "info")

        return redirect(url_for("team_tasks", user=filter_user, status=filter_status))

    tasks     = db.get_team_tasks(assigned_to=filter_user or None, status=filter_status or None)
    all_leads = db.get_all_leads()
    return render_template("team_tasks.html",
                           tasks=tasks, team=TEAM_DISPLAY,
                           all_leads=all_leads,
                           filter_user=filter_user,
                           filter_status=filter_status,
                           current_username=current_user.username)


@app.route("/team-tasks/<int:tid>/status", methods=["POST"])
@login_required
def team_task_quick_status(tid):
    db.update_team_task_status(tid, request.form.get("status", "Done"))
    redir = request.form.get("redirect", "team_tasks")
    if redir == "dashboard":
        return redirect(url_for("dashboard"))
    return redirect(url_for("team_tasks"))


# ── REMINDERS ──────────────────────────────────────────────────────────────────

@app.route("/reminders")
@login_required
def reminders():
    # Auto-fire overdue reminders
    due = db.get_due_reminders()
    fired = []
    for task in due:
        targets = list(TEAM_EMAILS.keys()) if task["assigned_to"] == "all" else [task["assigned_to"]]
        wa_msg = (f"⏰ Reminder: *{task['title']}*\n"
                  f"Due: {str(task.get('due_date',''))[:10] or 'N/A'}\n"
                  f"https://app.biminfrasolutions.in/team-tasks")
        for username in targets:
            _send_internal_email(
                username,
                f"Reminder: {task['title']}",
                _task_email_html(task, "reminder"),
            )
            _send_whatsapp_bot(username, wa_msg)
        db.mark_reminder_sent(task["id"])
        fired.append(task["title"])

    if fired:
        flash(f"Reminder emails sent for: {', '.join(fired)}", "info")

    # Show all upcoming + recent tasks with reminders
    upcoming = db.get_team_tasks()
    return render_template("reminders.html",
                           tasks=upcoming, team=TEAM_DISPLAY,
                           current_username=current_user.username,
                           now=datetime.utcnow())


@app.route("/reminders/send-now/<int:tid>", methods=["POST"])
@login_required
def send_reminder_now(tid):
    tasks = db.get_team_tasks()
    task  = next((t for t in tasks if t["id"] == tid), None)
    if task:
        targets = list(TEAM_EMAILS.keys()) if task["assigned_to"] == "all" else [task["assigned_to"]]
        wa_msg = (f"⏰ Reminder: *{task['title']}*\n"
                  f"Due: {str(task.get('due_date',''))[:10] or 'N/A'}\n"
                  f"https://app.biminfrasolutions.in/team-tasks")
        for username in targets:
            _send_internal_email(username, f"Reminder: {task['title']}", _task_email_html(task, "reminder"))
            _send_whatsapp_bot(username, wa_msg)
        db.mark_reminder_sent(tid)
        flash(f"Reminder sent for: {task['title']}", "success")
    return redirect(url_for("reminders"))


# ── DB BACKUP ─────────────────────────────────────────────────────────────────

@app.route("/admin/backup-db")
@login_required
def backup_db():
    """Download full DB as JSON (Kishan only)."""
    from flask import Response
    import json as _json
    if current_user.username != "kishan":
        return "Not authorized", 403
    tables = ["leads","email_logs","replies","tasks","team_tasks","responsibilities",
              "user_settings","expenses","income_entries","projects","invoices","invoice_items"]
    conn = db.get_db()
    c    = conn.cursor()
    backup = {}
    for t in tables:
        try:
            c.execute(f"SELECT * FROM {t}")
            backup[t] = db._fetchall(c.fetchall())
        except Exception:
            backup[t] = []
    conn.close()
    def _default(o):
        from datetime import date, datetime
        if isinstance(o, (date, datetime)):
            return str(o)
        return str(o)
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    data = _json.dumps(backup, default=_default, indent=2, ensure_ascii=False)
    return Response(data,
                    mimetype="application/json",
                    headers={"Content-Disposition": f"attachment; filename=bim_crm_backup_{ts}.json"})


# ── ONE-TIME MIGRATION ────────────────────────────────────────────────────────

@app.route("/admin/run-migration")
@login_required
def run_migration():
    """One-time: import migration_data.json into PostgreSQL. Only works for Kishan."""
    if current_user.username != "kishan":
        return "Not authorized", 403

    import json as _json
    migration_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "migration_data.json")
    if not os.path.exists(migration_file):
        return "migration_data.json not found", 404

    with open(migration_file) as f:
        data = _json.load(f)

    conn = db.get_db()
    c    = conn.cursor()

    leads_done = email_done = replies_done = tasks_done = 0

    # Check if already migrated
    c.execute("SELECT COUNT(*) AS cnt FROM leads")
    row = c.fetchone()
    existing = row["cnt"] if isinstance(row, dict) else (row[0] if row else 0)
    if existing > 0:
        conn.close()
        flash(f"Already has {existing} leads — migration skipped.", "warning")
        return redirect(url_for("leads"))

    # ── Leads ──────────────────────────────────────────────────────────────────
    for r in data.get("leads", []):
        try:
            c.execute(db._named("""
                INSERT INTO leads (
                    id, first_name, last_name, email, company, title, phone, website,
                    city, country, industry, status, priority_score, services_needed,
                    outsourcing_likelihood, pitch_angle, email_template, linkedin_url,
                    follow_up_stage, description, email_sequence_step, last_email_sent,
                    created_at, updated_at
                ) VALUES (
                    :id,:first_name,:last_name,:email,:company,:title,:phone,:website,
                    :city,:country,:industry,:status,:priority_score,:services_needed,
                    :outsourcing_likelihood,:pitch_angle,:email_template,:linkedin_url,
                    :follow_up_stage,:description,:email_sequence_step,:last_email_sent,
                    :created_at,:updated_at
                )
            """), r)
            leads_done += 1
        except Exception:
            pass
    conn.commit()
    if db._is_pg():
        c.execute("SELECT setval(pg_get_serial_sequence('leads','id'), (SELECT MAX(id) FROM leads))")
        conn.commit()

    # ── Email logs ─────────────────────────────────────────────────────────────
    for r in data.get("email_logs", []):
        try:
            # Normalise field names from SQLite export to current schema
            row = {
                "id":            r.get("id"),
                "lead_id":       r.get("lead_id"),
                "subject":       r.get("subject"),
                "body":          r.get("body"),
                "template_key":  r.get("template_key") or r.get("template_used"),
                "sequence_step": r.get("sequence_step"),
                "sent_at":       r.get("sent_at"),
                "opened_at":     r.get("opened_at") or r.get("opened"),
                "open_count":    r.get("open_count", 0),
                "clicked_at":    r.get("clicked_at") or r.get("clicked"),
                "bounced_at":    r.get("bounced_at") or r.get("bounced"),
                "bounce_reason": r.get("bounce_reason"),
                "status":        r.get("status"),
            }
            c.execute(db._named("""
                INSERT INTO email_logs (
                    id, lead_id, subject, body, template_key, sequence_step,
                    sent_at, opened_at, open_count, clicked_at, bounced_at, bounce_reason, status
                ) VALUES (
                    :id,:lead_id,:subject,:body,:template_key,:sequence_step,
                    :sent_at,:opened_at,:open_count,:clicked_at,:bounced_at,:bounce_reason,:status
                )
            """), row)
            email_done += 1
        except Exception:
            pass
    conn.commit()
    if db._is_pg() and email_done:
        c.execute("SELECT setval(pg_get_serial_sequence('email_logs','id'), (SELECT MAX(id) FROM email_logs))")
        conn.commit()

    # ── Replies ────────────────────────────────────────────────────────────────
    for r in data.get("replies", []):
        try:
            c.execute(db._named("""
                INSERT INTO replies (id, lead_id, from_email, subject, body, priority, status, source, received_at)
                VALUES (:id,:lead_id,:from_email,:subject,:body,:priority,:status,:source,:received_at)
            """), r)
            replies_done += 1
        except Exception:
            pass
    conn.commit()
    if db._is_pg() and replies_done:
        c.execute("SELECT setval(pg_get_serial_sequence('replies','id'), (SELECT MAX(id) FROM replies))")
        conn.commit()

    # ── Tasks ──────────────────────────────────────────────────────────────────
    for r in data.get("tasks", []):
        try:
            c.execute(db._named("""
                INSERT INTO tasks (id, lead_id, subject, due_date, priority, description, status, created_at)
                VALUES (:id,:lead_id,:subject,:due_date,:priority,:description,:status,:created_at)
            """), r)
            tasks_done += 1
        except Exception:
            pass
    conn.commit()
    if db._is_pg() and tasks_done:
        c.execute("SELECT setval(pg_get_serial_sequence('tasks','id'), (SELECT MAX(id) FROM tasks))")
        conn.commit()

    conn.close()
    flash(f"Migration complete: {leads_done} leads | {email_done} email logs | {replies_done} replies | {tasks_done} tasks", "success")
    return redirect(url_for("leads"))


# ── WHATSAPP ──────────────────────────────────────────────────────────────────

WA_TEMPLATES = {
    "intro": {
        "label": "Introduction",
        "body": "Hi {first_name}, I'm Kishan from BIM INFRASOLUTIONS LLP. We deliver BIM, Scan-to-BIM and coordination services for construction firms globally. Would you be open to a quick 10-min call this week? 🏗️"
    },
    "followup": {
        "label": "Follow-up",
        "body": "Hi {first_name}, following up on my earlier message. We've helped firms like {company} cut BIM delivery time by 40% with our offshore team. Happy to share a quick case study — shall I send it over?"
    },
    "portfolio": {
        "label": "Share Portfolio",
        "body": "Hi {first_name}, here is our BIM portfolio for your reference: https://drive.google.com/file/d/1AJ0_5XUJ5JxoHw3cXVJ7jFVh0LOZIkMk/view — 100+ projects across USA, Germany, Australia. Let me know if any project type is relevant to {company}."
    },
    "meeting": {
        "label": "Book Meeting",
        "body": "Hi {first_name}, would love to show you what we can do for {company}. You can book a 30-min call here: https://calendly.com/kishanbatavia9/30min — completely free, no obligation."
    },
    "custom": {
        "label": "Custom Message",
        "body": ""
    },
}


@app.route("/leads/<int:lead_id>/whatsapp", methods=["POST"])
@login_required
def send_whatsapp(lead_id):
    lead = db.get_lead(lead_id)
    if not lead:
        flash("Lead not found.", "danger")
        return redirect(url_for("leads"))

    phone   = request.form.get("phone", "").strip()
    message = request.form.get("message", "").strip()
    method  = request.form.get("method", "manual")  # manual | twilio

    if not phone:
        flash("No phone number provided.", "danger")
        return redirect(url_for("lead_detail", lead_id=lead_id))

    # Clean phone — remove spaces, dashes, keep +
    phone_clean = "".join(c for c in phone if c.isdigit() or c == "+")
    if not phone_clean.startswith("+"):
        phone_clean = "+" + phone_clean

    if method == "twilio":
        # Send via Twilio WhatsApp API
        account_sid = os.getenv("TWILIO_ACCOUNT_SID", "")
        auth_token  = os.getenv("TWILIO_AUTH_TOKEN", "")
        from_wa     = os.getenv("TWILIO_WHATSAPP_FROM", "")  # e.g. whatsapp:+14155238886
        if not account_sid or not auth_token or not from_wa:
            flash("Twilio credentials not set. Add TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_WHATSAPP_FROM to environment.", "danger")
            return redirect(url_for("lead_detail", lead_id=lead_id))
        try:
            import requests as req
            r = req.post(
                f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json",
                auth=(account_sid, auth_token),
                data={
                    "From": from_wa,
                    "To"  : f"whatsapp:{phone_clean}",
                    "Body": message,
                },
                timeout=15,
            )
            if r.status_code in (200, 201):
                db.log_whatsapp(lead_id, phone_clean, message, current_user.username, "twilio", "sent")
                if lead.get("status") == "New":
                    db.update_lead_status(lead_id, "Contacted")
                flash(f"WhatsApp sent to {phone_clean} via Twilio!", "success")
            else:
                err = r.json().get("message", r.text[:100])
                flash(f"Twilio error: {err}", "danger")
        except Exception as e:
            flash(f"WhatsApp send failed: {e}", "danger")
    else:
        # Manual — just log it (user opened wa.me link themselves)
        db.log_whatsapp(lead_id, phone_clean, message, current_user.username, "manual", "sent")
        if lead.get("status") == "New":
            db.update_lead_status(lead_id, "Contacted")
        flash("WhatsApp message logged.", "success")

    return redirect(url_for("lead_detail", lead_id=lead_id))


@app.route("/api/wa-templates")
@login_required
def wa_templates_api():
    return jsonify(WA_TEMPLATES)


# ── EXPENSES ──────────────────────────────────────────────────────────────────

@app.route("/expenses", methods=["GET", "POST"])
@login_required
def expenses():
    filter_partner = request.args.get("partner", "")
    filter_month   = request.args.get("month", "")
    filter_type    = request.args.get("type", "")
    filter_project = request.args.get("project", "")

    if request.method == "POST":
        action = request.form.get("action", "create")
        if action == "create":
            raw_date = request.form.get("date", "")
            try:
                month_year = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%b %Y")
            except Exception:
                month_year = ""
            data = {
                "date"        : raw_date,
                "month_year"  : month_year,
                "partner"     : request.form.get("partner", ""),
                "expense_type": request.form.get("expense_type", "Common"),
                "project_name": request.form.get("project_name", "") or None,
                "category"    : request.form.get("category", ""),
                "description" : request.form.get("description", "").strip(),
                "amount"      : float(request.form.get("amount", 0) or 0),
                "created_by"  : current_user.username,
            }
            db.create_expense(data)
            flash("Expense added.", "success")
        elif action == "edit":
            eid = int(request.form.get("expense_id"))
            raw_date = request.form.get("date", "")
            try:
                month_year = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%b %Y")
            except Exception:
                month_year = ""
            db.update_expense(eid, {
                "date"        : raw_date,
                "month_year"  : month_year,
                "partner"     : request.form.get("partner", ""),
                "expense_type": request.form.get("expense_type", "Common"),
                "project_name": request.form.get("project_name", "") or None,
                "category"    : request.form.get("category", ""),
                "description" : request.form.get("description", "").strip(),
                "amount"      : float(request.form.get("amount", 0) or 0),
            })
            flash("Expense updated.", "success")
        elif action == "delete":
            db.delete_expense(int(request.form.get("expense_id")))
            flash("Expense deleted.", "info")
        return redirect(url_for("expenses", partner=filter_partner, month=filter_month,
                                type=filter_type, project=filter_project))

    rows = db.get_expenses(
        partner=filter_partner or None,
        month_year=filter_month or None,
        expense_type=filter_type or None,
        project=filter_project or None,
    )
    summary = db.get_expense_summary()
    return render_template("expenses.html",
                           expenses=rows,
                           summary=summary,
                           categories=db.EXPENSE_CATEGORIES,
                           partners=db.PARTNERS,
                           projects=db.PROJECTS,
                           income_categories=db.INCOME_CATEGORIES,
                           filter_partner=filter_partner,
                           filter_month=filter_month,
                           filter_type=filter_type,
                           filter_project=filter_project,
                           now=datetime.now())


@app.route("/expenses/export-excel")
@login_required
def expenses_export_excel():
    """Download project expenses as monthly Excel sheet."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import Response

    filter_project = request.args.get("project", "") or None
    filter_month   = request.args.get("month", "") or None

    rows = db.get_expenses(
        partner=None,
        month_year=filter_month or None,
        expense_type="Project",
        project=filter_project or None,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    title_parts = ["Project Expenses"]
    if filter_project:
        title_parts.append(filter_project)
    if filter_month:
        title_parts.append(filter_month)
    ws.title = "Expenses"

    # ── Header Row ──
    gold_fill   = PatternFill("solid", fgColor="D4A017")
    dark_fill   = PatternFill("solid", fgColor="1A1A1A")
    alt_fill    = PatternFill("solid", fgColor="F5F5F5")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = " | ".join(title_parts)
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill  = dark_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Date", "Partner", "Project", "Category", "Description", "Amount (₹)", "Added By"]
    col_widths = [5, 14, 16, 22, 20, 36, 16, 14]
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font  = Font(name="Calibri", bold=True, size=11, color="000000")
        cell.fill  = gold_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20

    # Data rows
    total = 0.0
    for idx, e in enumerate(rows, start=1):
        r = idx + 2
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        vals = [
            idx,
            str(e.get("date", ""))[:10],
            e.get("partner", ""),
            e.get("project_name", "") or "",
            e.get("category", ""),
            e.get("description", "") or "",
            float(e.get("amount", 0) or 0),
            e.get("created_by", ""),
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font   = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.fill   = fill
            if ci == 7:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif ci == 1:
                cell.alignment = Alignment(horizontal="center")
        total += float(e.get("amount", 0) or 0)

    # Total row
    total_row = len(rows) + 3
    ws.merge_cells(f"A{total_row}:F{total_row}")
    tc = ws[f"A{total_row}"]
    tc.value = "TOTAL"
    tc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    tc.fill  = dark_fill
    tc.alignment = Alignment(horizontal="right")
    tc.border = thin_border
    amt_cell = ws.cell(row=total_row, column=7, value=total)
    amt_cell.font   = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    amt_cell.fill   = dark_fill
    amt_cell.number_format = '#,##0.00'
    amt_cell.alignment = Alignment(horizontal="right")
    amt_cell.border = thin_border
    ws.cell(row=total_row, column=8).fill   = dark_fill
    ws.cell(row=total_row, column=8).border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname_parts = ["project_expenses"]
    if filter_project:
        fname_parts.append(filter_project.replace(" ", "_"))
    if filter_month:
        fname_parts.append(filter_month.replace(" ", "_"))
    filename = "_".join(fname_parts) + ".xlsx"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/income", methods=["GET", "POST"])
@login_required
def income():
    filter_month    = request.args.get("month", "")
    filter_category = request.args.get("category", "")

    if request.method == "POST":
        action = request.form.get("action", "create")
        if action == "create":
            raw_date = request.form.get("date", "")
            try:
                month_year = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%b %Y")
            except Exception:
                month_year = ""
            db.create_income({
                "date"           : raw_date,
                "month_year"     : month_year,
                "client_source"  : request.form.get("client_source", "").strip(),
                "income_category": request.form.get("income_category", ""),
                "description"    : request.form.get("description", "").strip(),
                "amount"         : float(request.form.get("amount", 0) or 0),
                "created_by"     : current_user.username,
            })
            flash("Income entry added.", "success")
        elif action == "edit":
            iid = int(request.form.get("income_id"))
            raw_date = request.form.get("date", "")
            try:
                month_year = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%b %Y")
            except Exception:
                month_year = ""
            db.update_income(iid, {
                "date"           : raw_date,
                "month_year"     : month_year,
                "client_source"  : request.form.get("client_source", "").strip(),
                "income_category": request.form.get("income_category", ""),
                "description"    : request.form.get("description", "").strip(),
                "amount"         : float(request.form.get("amount", 0) or 0),
            })
            flash("Income updated.", "success")
        elif action == "delete":
            db.delete_income(int(request.form.get("income_id")))
            flash("Income entry deleted.", "info")
        return redirect(url_for("income", month=filter_month, category=filter_category))

    rows = db.get_income(month_year=filter_month or None, category=filter_category or None)
    summary = db.get_expense_summary()
    return render_template("income.html",
                           income_entries=rows,
                           summary=summary,
                           income_categories=db.INCOME_CATEGORIES,
                           filter_month=filter_month,
                           filter_category=filter_category,
                           now=datetime.now())


# ── PROJECTS ──────────────────────────────────────────────────────────────────

@app.route("/projects", methods=["GET", "POST"])
@login_required
def projects():
    filter_status = request.args.get("status", "")
    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            db.create_project({
                "name"          : request.form.get("name", "").strip(),
                "client_name"   : request.form.get("client_name", "").strip(),
                "client_address": request.form.get("client_address", "").strip(),
                "client_gstin"  : request.form.get("client_gstin", "").strip(),
                "start_date"    : request.form.get("start_date") or None,
                "end_date"      : request.form.get("end_date") or None,
                "status"        : request.form.get("status", "Active"),
                "total_value"   : float(request.form.get("total_value") or 0),
                "description"   : request.form.get("description", "").strip(),
                "lead_id"       : request.form.get("lead_id") or None,
                "created_by"    : current_user.username,
            })
            flash("Project created.", "success")
        elif action == "edit":
            pid = int(request.form.get("project_id"))
            db.update_project(pid, {
                "name"          : request.form.get("name", "").strip(),
                "client_name"   : request.form.get("client_name", "").strip(),
                "client_address": request.form.get("client_address", "").strip(),
                "client_gstin"  : request.form.get("client_gstin", "").strip(),
                "start_date"    : request.form.get("start_date") or None,
                "end_date"      : request.form.get("end_date") or None,
                "status"        : request.form.get("status", "Active"),
                "total_value"   : float(request.form.get("total_value") or 0),
                "description"   : request.form.get("description", "").strip(),
            })
            flash("Project updated.", "success")
        elif action == "delete":
            db.delete_project(int(request.form.get("project_id")))
            flash("Project deleted.", "info")
        return redirect(url_for("projects", status=filter_status))

    all_projects = db.get_projects(status=filter_status or None)
    return render_template("projects.html",
                           projects=all_projects,
                           filter_status=filter_status,
                           now=datetime.now())


@app.route("/leads/<int:lead_id>/to-project", methods=["GET", "POST"])
@login_required
def lead_to_project(lead_id):
    lead = db.get_lead(lead_id)
    if not lead:
        flash("Lead not found.", "danger")
        return redirect(url_for("leads"))
    if request.method == "POST":
        pid = db.create_project({
            "name"          : request.form.get("name", lead.get("company", "")).strip(),
            "client_name"   : request.form.get("client_name", lead.get("company", "")).strip(),
            "client_address": request.form.get("client_address", "").strip(),
            "client_gstin"  : request.form.get("client_gstin", "").strip(),
            "start_date"    : request.form.get("start_date") or None,
            "end_date"      : request.form.get("end_date") or None,
            "status"        : "Active",
            "total_value"   : float(request.form.get("total_value") or 0),
            "description"   : request.form.get("description", "").strip(),
            "lead_id"       : lead_id,
            "created_by"    : current_user.username,
        })
        # Mark lead as converted
        db.update_lead_status(lead_id, "Client")
        flash(f"Lead converted to project #{pid}.", "success")
        return redirect(url_for("projects"))
    return render_template("lead_to_project.html", lead=lead, now=datetime.now())


# ── INVOICES ──────────────────────────────────────────────────────────────────

@app.route("/invoices")
@login_required
def invoices():
    filter_status  = request.args.get("status", "")
    filter_project = request.args.get("project_id", "")
    all_invoices = db.get_invoices(
        status=filter_status or None,
        project_id=int(filter_project) if filter_project else None
    )
    summary  = db.get_invoice_summary()
    projects = db.get_projects()
    return render_template("invoices.html",
                           invoices=all_invoices,
                           summary=summary,
                           projects=projects,
                           filter_status=filter_status,
                           filter_project=filter_project,
                           now=datetime.now())


@app.route("/invoices/new", methods=["GET", "POST"])
@login_required
def invoice_new():
    if request.method == "POST":
        return _save_invoice(None)
    date_default = datetime.now().strftime("%Y-%m-%d")
    inv_no = db.next_invoice_number(date_default)
    projects = db.get_projects()
    return render_template("invoice_form.html", invoice=None, items=[],
                           inv_no=inv_no, date_default=date_default,
                           projects=projects, now=datetime.now())


@app.route("/invoices/<int:invoice_id>/edit", methods=["GET", "POST"])
@login_required
def invoice_edit(invoice_id):
    invoice = db.get_invoice(invoice_id)
    if not invoice:
        flash("Invoice not found.", "danger")
        return redirect(url_for("invoices"))
    if request.method == "POST":
        return _save_invoice(invoice_id)
    items    = db.get_invoice_items(invoice_id)
    projects = db.get_projects()
    return render_template("invoice_form.html", invoice=invoice, items=items,
                           inv_no=invoice["invoice_no"],
                           date_default=invoice.get("date", ""),
                           projects=projects, now=datetime.now())


@app.route("/invoices/<int:invoice_id>")
@login_required
def invoice_view(invoice_id):
    invoice = db.get_invoice(invoice_id)
    if not invoice:
        flash("Invoice not found.", "danger")
        return redirect(url_for("invoices"))
    items = db.get_invoice_items(invoice_id)
    return render_template("invoice_view.html", invoice=invoice, items=items, now=datetime.now())


@app.route("/invoices/<int:invoice_id>/status", methods=["POST"])
@login_required
def invoice_status(invoice_id):
    db.update_invoice_status(invoice_id, request.form.get("status", "Draft"))
    flash("Invoice status updated.", "success")
    return redirect(url_for("invoice_view", invoice_id=invoice_id))


@app.route("/invoices/<int:invoice_id>/delete", methods=["POST"])
@login_required
def invoice_delete(invoice_id):
    db.delete_invoice(invoice_id)
    flash("Invoice deleted.", "info")
    return redirect(url_for("invoices"))


@app.route("/invoices/<int:invoice_id>/download/docx")
@login_required
def invoice_download_docx(invoice_id):
    from flask import Response
    import invoice_generator as ig
    invoice = db.get_invoice(invoice_id)
    items   = db.get_invoice_items(invoice_id)
    if not invoice:
        return "Not found", 404
    data = ig.generate_docx(invoice, items)
    fname = f"Invoice_{invoice['invoice_no'].replace('/', '_')}.docx"
    return Response(data,
                    mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/invoices/<int:invoice_id>/download/pdf")
@login_required
def invoice_download_pdf(invoice_id):
    from flask import Response
    import invoice_generator as ig
    invoice = db.get_invoice(invoice_id)
    items   = db.get_invoice_items(invoice_id)
    if not invoice:
        return "Not found", 404
    data = ig.generate_pdf(invoice, items)
    fname = f"Invoice_{invoice['invoice_no'].replace('/', '_')}.pdf"
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/invoices/<int:invoice_id>/send-email", methods=["POST"])
@login_required
def invoice_send_email(invoice_id):
    """Send invoice to client and/or accountant by email."""
    from zoho_mail import mail_client
    invoice = db.get_invoice(invoice_id)
    items   = db.get_invoice_items(invoice_id)
    if not invoice:
        flash("Invoice not found.", "danger")
        return redirect(url_for("invoices"))

    client_email    = request.form.get("client_email", "").strip()
    accountant_email = request.form.get("accountant_email", "").strip()
    user_cfg        = db.get_user_settings(current_user.username)

    # Build HTML body
    items_rows = "".join(
        f"<tr><td style='padding:8px;border:1px solid #333;'>{it['description']}"
        f"{'<br><small style=\"color:#aaa\">(SAC: ' + it['sac_code'] + ')</small>' if it.get('sac_code') else ''}</td>"
        f"<td style='padding:8px;border:1px solid #333;text-align:center;'>{it['unit']}</td>"
        f"<td style='padding:8px;border:1px solid #333;text-align:right;'>₹{float(it['amount'] or 0):,.0f}</td></tr>"
        for it in items
    )
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:640px;margin:0 auto;background:#0e0e0e;color:#e8e8e8;padding:24px;border-radius:8px;">
      <h2 style="color:#D4A017;text-align:center;margin:0 0 4px;">BIM INFRASOLUTIONS LLP</h2>
      <p style="text-align:center;color:#888;font-size:12px;margin:0 0 20px;">GSTIN: 24AAUFB9689E1ZS | LLPIN: AAP-1096</p>
      <hr style="border-color:#333;">
      <h3 style="text-align:center;color:#fff;">Tax Invoice — {invoice['invoice_no']}</h3>
      <p style="color:#aaa;font-size:13px;"><strong>Date:</strong> {invoice['date']}</p>
      <p style="color:#e8e8e8;font-size:13px;"><strong>To:</strong> {invoice['client_name']}<br>
      {'<small>' + invoice.get('client_address','') + '</small><br>' if invoice.get('client_address') else ''}
      {'GSTIN: ' + invoice['client_gstin'] if invoice.get('client_gstin') else ''}</p>
      <table style="width:100%;border-collapse:collapse;font-size:13px;margin:16px 0;">
        <thead>
          <tr style="background:#1f1f1f;">
            <th style="padding:8px;border:1px solid #333;text-align:left;">Particulars</th>
            <th style="padding:8px;border:1px solid #333;text-align:center;">Unit</th>
            <th style="padding:8px;border:1px solid #333;text-align:right;">Amount (INR)</th>
          </tr>
        </thead>
        <tbody>
          {items_rows}
          <tr><td style="padding:8px;border:1px solid #333;">GST {int(invoice.get('gst_rate',18))}%</td>
              <td style="border:1px solid #333;"></td>
              <td style="padding:8px;border:1px solid #333;text-align:right;">₹{float(invoice.get('gst_amount',0)):,.0f}</td></tr>
          <tr style="background:#1a1a1a;">
            <td colspan="2" style="padding:8px;border:1px solid #333;font-weight:bold;color:#D4A017;">Total Bill</td>
            <td style="padding:8px;border:1px solid #333;text-align:right;font-weight:bold;color:#D4A017;font-size:15px;">
              ₹{float(invoice.get('total',0)):,.0f} INR</td>
          </tr>
        </tbody>
      </table>
      <p style="font-size:12px;color:#888;"><strong>Payment within 14 days.</strong> HDFC Bank | IFSC: HDFC0003905 | A/C: 50200041261501</p>
      <hr style="border-color:#333;">
      <p style="font-size:11px;color:#555;text-align:center;">
        302, Shahibaug Greens, Ahmedabad · info@biminfrasolutions.in · www.biminfrasolutions.in
      </p>
    </div>"""

    sent_to = []
    failed  = []

    if client_email:
        ok = mail_client.send_email(
            to_address=client_email,
            subject=f"Invoice {invoice['invoice_no']} from BIM Infrasolutions LLP",
            html_body=html,
            user_settings=user_cfg,
        )
        if ok: sent_to.append(f"client ({client_email})")
        else:  failed.append(client_email)

    if accountant_email:
        acct_html = html + f"""
        <div style="background:#1f1f1f;border-radius:6px;padding:12px;margin-top:16px;font-size:12px;color:#aaa;">
          <strong style="color:#D4A017;">Accountant Note:</strong><br>
          Invoice <strong>{invoice['invoice_no']}</strong> generated by {current_user.display} on {invoice['date']}.
          Please record payment once received.
        </div>"""
        ok = mail_client.send_email(
            to_address=accountant_email,
            subject=f"[Internal] Invoice {invoice['invoice_no']} generated — {invoice['client_name']}",
            html_body=acct_html,
            user_settings=user_cfg,
        )
        if ok: sent_to.append(f"accountant ({accountant_email})")
        else:  failed.append(accountant_email)

    if sent_to:
        db.update_invoice_status(invoice_id, "Sent")
        flash(f"Invoice sent to: {', '.join(sent_to)}. Status updated to Sent.", "success")
    if failed:
        flash(f"Failed to send to: {', '.join(failed)}", "danger")
    return redirect(url_for("invoice_view", invoice_id=invoice_id))


def _save_invoice(invoice_id):
    """Parse invoice form and create/update invoice + items."""
    date_str    = request.form.get("date", "")
    inv_no      = request.form.get("invoice_no", "").strip()
    project_id  = request.form.get("project_id") or None
    gst_rate    = float(request.form.get("gst_rate", 18) or 18)

    # Parse line items
    descriptions = request.form.getlist("item_description[]")
    sac_codes    = request.form.getlist("item_sac[]")
    units        = request.form.getlist("item_unit[]")
    rates        = request.form.getlist("item_rate[]")

    items = []
    subtotal = 0.0
    for i, desc in enumerate(descriptions):
        if not desc.strip():
            continue
        rate = float(rates[i] if i < len(rates) else 0 or 0)
        unit = int(units[i] if i < len(units) else 1 or 1)
        amt  = rate * unit
        subtotal += amt
        items.append({
            "description": desc.strip(),
            "sac_code"   : sac_codes[i] if i < len(sac_codes) else "",
            "unit"       : unit,
            "rate"       : rate,
            "amount"     : amt,
        })

    gst_amount = round(subtotal * gst_rate / 100, 2)
    total      = round(subtotal + gst_amount, 2)

    data = {
        "invoice_no"     : inv_no,
        "date"           : date_str,
        "project_id"     : int(project_id) if project_id else None,
        "client_name"    : request.form.get("client_name", "").strip(),
        "client_address" : request.form.get("client_address", "").strip(),
        "client_gstin"   : request.form.get("client_gstin", "").strip(),
        "lut_number"     : request.form.get("lut_number", "").strip(),
        "gst_rate"       : gst_rate,
        "subtotal"       : subtotal,
        "gst_amount"     : gst_amount,
        "total"          : total,
        "notes"          : request.form.get("notes", "").strip(),
        "status"         : request.form.get("status", "Draft"),
        "created_by"     : current_user.username,
    }

    if invoice_id:
        db.update_invoice(invoice_id, data, items)
        flash("Invoice updated.", "success")
        return redirect(url_for("invoice_view", invoice_id=invoice_id))
    else:
        new_id = db.create_invoice(data, items)
        flash("Invoice created.", "success")
        return redirect(url_for("invoice_view", invoice_id=new_id))


# ── ENTRY POINT ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  BIM Infra Solutions — Custom CRM")
    print("=" * 55)
    print("\n  Open your browser and go to:")
    print("  http://localhost:5000")
    print("\n  Press Ctrl+C to stop the server\n")
    app.run(debug=True, host="0.0.0.0", port=5000)
