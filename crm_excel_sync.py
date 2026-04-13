"""
BIM Infra Solutions — CRM ↔ Excel Two-Way Sync Engine
======================================================
File  : crm_excel_sync.py
Place : C:\\Users\\Kishan\\BIM_CRM\\crm_excel_sync.py

WHAT THIS DOES
--------------
1. EXPORT  → Reads bim_crm.db  → Writes BIM_CRM_Sync.xlsx (all tables, live stats)
2. IMPORT  → Reads Excel sheet "Lead Database" → Inserts NEW leads into DB (skips duplicates)
3. PUSH-STATUS → Reads Excel "CRM Status Tracker" changes → Updates DB statuses
4. MAIL-SYNC → Polls Zoho Mail inbox → Finds replies from lead emails → Auto-updates DB
5. FULL-SYNC → Runs all of the above in sequence

USAGE
-----
  python crm_excel_sync.py export          # DB → Excel
  python crm_excel_sync.py import          # Excel → DB (new leads only)
  python crm_excel_sync.py push-status     # Excel status changes → DB
  python crm_excel_sync.py mail-sync       # Zoho inbox → DB updates
  python crm_excel_sync.py full-sync       # Everything at once
  python crm_excel_sync.py watch           # Auto full-sync every 30 min

EXCEL FILE: BIM_CRM_Sync.xlsx  (created in same folder as this script)

SHEETS
------
  "Lead Database"       — CRM import-compatible (matches app.py exactly)
  "CRM Status Tracker"  — Live DB view: all fields incl. status, sequence step
  "Email Activity"      — email_logs table (opens, clicks, sent)
  "Tasks"               — tasks table (pending + completed)
  "Dashboard"           — Summary stats, KPI gauges, status breakdown
  "Zoho Replies"        — Replies detected from Zoho Mail inbox
"""

import os, sys, json, time, logging, sqlite3
import requests
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler("crm_sync.log"), logging.StreamHandler()],
)
log = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
DB_PATH    = BASE_DIR / "bim_crm.db"
EXCEL_PATH = BASE_DIR / "BIM_CRM_Sync.xlsx"

# ── Zoho Config ────────────────────────────────────────────────────────────────
ZOHO_CLIENT_ID     = os.getenv("ZOHO_CLIENT_ID")
ZOHO_CLIENT_SECRET = os.getenv("ZOHO_CLIENT_SECRET")
ZOHO_REFRESH_TOKEN = os.getenv("ZOHO_REFRESH_TOKEN")
ZOHO_DC            = os.getenv("ZOHO_DC", "in")
ZOHO_ACCOUNT_ID    = os.getenv("ZOHO_MAIL_ACCOUNT_ID", "")
SENDER_EMAIL       = "kishan@biminfrasolutions.in"

# ═══════════════════════════════════════════════════════════════════════════════
#  DB HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def fetch_all_leads():
    conn = get_db()
    rows = conn.execute("SELECT * FROM leads ORDER BY priority_score DESC, created_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def fetch_email_logs():
    conn = get_db()
    rows = conn.execute("""
        SELECT e.id, e.lead_id, l.first_name||' '||COALESCE(l.last_name,'') AS full_name,
               l.company, l.email AS lead_email,
               e.subject, e.template_used, e.sequence_step,
               e.sent_at, e.status, e.opened, e.clicked, e.open_count
        FROM email_logs e
        JOIN leads l ON e.lead_id = l.id
        ORDER BY e.sent_at DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def fetch_tasks():
    conn = get_db()
    rows = conn.execute("""
        SELECT t.id, t.lead_id, l.first_name||' '||COALESCE(l.last_name,'') AS full_name,
               l.company, l.email AS lead_email,
               t.subject, t.due_date, t.status, t.priority, t.description, t.created_at
        FROM tasks t
        JOIN leads l ON t.lead_id = l.id
        ORDER BY t.due_date ASC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def fetch_stats():
    conn = get_db()
    s = {
        "total_leads"   : conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0],
        "hot_leads"     : conn.execute("SELECT COUNT(*) FROM leads WHERE status='Hot'").fetchone()[0],
        "new_leads"     : conn.execute("SELECT COUNT(*) FROM leads WHERE status='New'").fetchone()[0],
        "contacted"     : conn.execute("SELECT COUNT(*) FROM leads WHERE status='Contacted'").fetchone()[0],
        "replied"       : conn.execute("SELECT COUNT(*) FROM leads WHERE status IN ('Replied','Engaged — Reply Received')").fetchone()[0],
        "pilot_sent"    : conn.execute("SELECT COUNT(*) FROM leads WHERE follow_up_stage LIKE '%Pilot%'").fetchone()[0],
        "emails_sent"   : conn.execute("SELECT COUNT(*) FROM email_logs").fetchone()[0],
        "emails_opened" : conn.execute("SELECT COUNT(*) FROM email_logs WHERE opened=1").fetchone()[0],
        "emails_clicked": conn.execute("SELECT COUNT(*) FROM email_logs WHERE clicked=1").fetchone()[0],
        "pending_tasks" : conn.execute("SELECT COUNT(*) FROM tasks WHERE status!='Completed'").fetchone()[0],
        "status_counts" : {r[0]:r[1] for r in conn.execute("SELECT status, COUNT(*) FROM leads GROUP BY status").fetchall()},
        "template_counts": {r[0]:r[1] for r in conn.execute("SELECT email_template, COUNT(*) FROM leads GROUP BY email_template").fetchall()},
        "country_counts" : {r[0]:r[1] for r in conn.execute("SELECT country, COUNT(*) FROM leads GROUP BY country ORDER BY COUNT(*) DESC LIMIT 20").fetchall()},
        "seq_step_counts": {r[0]:r[1] for r in conn.execute("SELECT email_sequence_step, COUNT(*) FROM leads GROUP BY email_sequence_step").fetchall()},
    }
    conn.close()
    return s

def upsert_lead(data: dict) -> str:
    """Insert new lead; if email exists, skip. Returns 'inserted'|'skipped'."""
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO leads (
                first_name, last_name, email, company, title, phone, website,
                city, country, industry, status, priority_score, services_needed,
                outsourcing_likelihood, pitch_angle, email_template, linkedin_url,
                follow_up_stage, description
            ) VALUES (
                :first_name,:last_name,:email,:company,:title,:phone,:website,
                :city,:country,:industry,:status,:priority_score,:services_needed,
                :outsourcing_likelihood,:pitch_angle,:email_template,:linkedin_url,
                :follow_up_stage,:description
            )
        """, data)
        conn.commit()
        conn.close()
        return "inserted"
    except sqlite3.IntegrityError:
        conn.close()
        return "skipped"

def update_lead_field(email: str, field: str, value):
    """Update a single field on a lead identified by email."""
    conn = get_db()
    conn.execute(f"UPDATE leads SET {field}=?, updated_at=? WHERE email=?",
                 (value, datetime.utcnow().isoformat(), email))
    conn.commit()
    conn.close()

def log_reply(lead_id: int, zoho_email_id: str, subject: str, received_at: str):
    """Log a detected reply in email_logs and update lead status."""
    conn = get_db()
    # Check if already logged
    existing = conn.execute(
        "SELECT id FROM email_logs WHERE lead_id=? AND subject=? AND status='reply'",
        (lead_id, subject)
    ).fetchone()
    if not existing:
        conn.execute("""
            INSERT INTO email_logs (lead_id, subject, body, template_used, sequence_step, status, sent_at)
            VALUES (?,?,?,?,?,?,?)
        """, (lead_id, f"REPLY: {subject}", f"Reply detected via Zoho Mail. ID:{zoho_email_id}",
              "REPLY", -1, "reply", received_at))
        conn.execute(
            "UPDATE leads SET status='Replied', follow_up_stage='Reply Received — Action Required', updated_at=? WHERE id=?",
            (datetime.utcnow().isoformat(), lead_id)
        )
        conn.commit()
        log.info("Reply logged for lead ID %s — Subject: %s", lead_id, subject)
    conn.close()

# ═══════════════════════════════════════════════════════════════════════════════
#  ZOHO MAIL — Token + Reply Detection
# ═══════════════════════════════════════════════════════════════════════════════
class ZohoMailSync:
    def __init__(self):
        self._token = None
        self._expiry = 0

    def get_token(self) -> str:
        if time.time() > self._expiry - 60:
            self._refresh()
        return self._token

    def _refresh(self):
        r = requests.post(
            f"https://accounts.zoho.{ZOHO_DC}/oauth/v2/token",
            params={"grant_type":"refresh_token","client_id":ZOHO_CLIENT_ID,
                    "client_secret":ZOHO_CLIENT_SECRET,"refresh_token":ZOHO_REFRESH_TOKEN},
            timeout=10
        )
        r.raise_for_status()
        data = r.json()
        if "access_token" not in data:
            raise Exception(f"Token refresh failed: {data}")
        self._token  = data["access_token"]
        self._expiry = time.time() + data.get("expires_in", 3600)
        log.info("Zoho token refreshed.")

    def _headers(self):
        return {"Authorization": f"Zoho-oauthtoken {self.get_token()}"}

    def get_account_id(self) -> str:
        """Resolve numeric account ID from Zoho Mail API."""
        if ZOHO_ACCOUNT_ID and ZOHO_ACCOUNT_ID.isdigit():
            return ZOHO_ACCOUNT_ID
        r = requests.get(f"https://mail.zoho.{ZOHO_DC}/api/accounts",
                         headers=self._headers(), timeout=10)
        r.raise_for_status()
        accounts = r.json().get("data", [])
        if accounts:
            aid = str(accounts[0].get("accountId", ""))
            log.info("Resolved Zoho Account ID: %s", aid)
            return aid
        raise Exception("No Zoho Mail accounts found.")

    def fetch_inbox(self, limit: int = 50) -> list:
        """Fetch recent inbox messages."""
        try:
            acct_id = self.get_account_id()
            r = requests.get(
                f"https://mail.zoho.{ZOHO_DC}/api/accounts/{acct_id}/messages/view",
                headers=self._headers(),
                params={"folderId": "INBOX", "limit": limit, "sortorder": "false"},
                timeout=15,
            )
            if r.status_code == 200:
                return r.json().get("data", [])
            log.warning("Inbox fetch returned %s: %s", r.status_code, r.text[:200])
            return []
        except Exception as e:
            log.error("Inbox fetch failed: %s", e)
            return []

    def check_replies(self) -> list:
        """
        Scan inbox for emails FROM lead email addresses.
        Returns list of detected replies: [{lead_id, email, subject, received_at, zoho_id}]
        """
        # Build email → lead_id map from DB
        conn = get_db()
        email_map = {r[0]: r[1] for r in
                     conn.execute("SELECT email, id FROM leads").fetchall()}
        conn.close()

        if not email_map:
            log.info("No leads in DB yet — skipping reply check.")
            return []

        messages = self.fetch_inbox(limit=100)
        replies  = []

        for msg in messages:
            sender = msg.get("fromAddress", "").lower().strip()
            # Zoho inbox message structure: fromAddress might be "Name <email>" format
            if "<" in sender:
                sender = sender.split("<")[1].rstrip(">").strip()

            if sender in email_map:
                lead_id  = email_map[sender]
                subject  = msg.get("subject", "(no subject)")
                recv_at  = msg.get("receivedTime", datetime.utcnow().isoformat())
                zoho_id  = str(msg.get("messageId", ""))

                log_reply(lead_id, zoho_id, subject, recv_at)
                replies.append({
                    "lead_id"    : lead_id,
                    "email"      : sender,
                    "subject"    : subject,
                    "received_at": recv_at,
                })
                log.info("REPLY DETECTED: %s — %s", sender, subject)

        log.info("Reply check complete. %d replies found.", len(replies))
        return replies

zoho_sync = ZohoMailSync()

# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def build_excel():
    """
    Build / refresh BIM_CRM_Sync.xlsx from the live SQLite database.
    Creates 6 sheets perfectly matched to the CRM schema.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        return

    leads      = fetch_all_leads()
    email_logs = fetch_email_logs()
    tasks      = fetch_tasks()
    stats      = fetch_stats()
    now_str    = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────────────────────────
    DARK     = "1B3A6B"; GOLD="D4A017"; LB="E8F0FE"; WHITE="FFFFFF"
    GREEN    = "00875A"; ORANGE="FF6B35"; RED="D93025"; GREY="F5F5F5"
    DARK_TXT = "1A1A2E"; TEAL="006D75"; PURPLE="5C0099"

    STATUS_COLORS = {
        "New"       : ("E3F2FD","1565C0"),
        "Contacted" : ("FFF9C4","F57F17"),
        "Hot"       : ("FFCDD2","C62828"),
        "Engaged — Reply Received": ("C8E6C9","2E7D32"),
        "Replied"   : ("C8E6C9","2E7D32"),
        "Cold"      : ("ECEFF1","546E7A"),
        "Pilot Sent": ("F3E5F5","7B1FA2"),
        "Converted" : ("DCEDC8","33691E"),
        "Invalid"   : ("FFEBEE","B71C1C"),
        "Unsubscribed":("FFEBEE","B71C1C"),
    }

    def hf(sz=10, bold=True, color=WHITE):
        return Font(name="Arial", size=sz, bold=bold, color=color)
    def cf(sz=9, bold=False, color=DARK_TXT):
        return Font(name="Arial", size=sz, bold=bold, color=color)
    def fp(hex_c):
        return PatternFill("solid", fgColor=hex_c)
    def tb():
        s=Side(style="thin",color="DDDDDD"); return Border(left=s,right=s,top=s,bottom=s)
    def ca():
        return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def la():
        return Alignment(horizontal="left",vertical="center",wrap_text=True)

    def write_header_row(ws, row_num, headers_widths, bg_color=DARK):
        for ci,(h,w) in enumerate(headers_widths, 1):
            c = ws.cell(row=row_num, column=ci, value=h)
            c.font=hf(10); c.fill=fp(bg_color); c.alignment=ca(); c.border=tb()
            ws.column_dimensions[get_column_letter(ci)].width=w

    def write_title(ws, text, n_cols, bg=DARK):
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        tc=ws["A1"]; tc.value=text
        tc.font=Font(name="Arial",size=12,bold=True,color=WHITE)
        tc.fill=fp(bg); tc.alignment=ca()
        ws.row_dimensions[1].height=26

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 1: "Lead Database"  (CRM import-compatible — matches app.py exactly)
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Lead Database"
    ws1.freeze_panes = "A3"

    write_title(ws1,
        f"BIM Infra Solutions — CRM Lead Import Sheet  |  {len(leads)} leads  |  Synced: {now_str}  |  Import via CRM → /leads/import",
        15, DARK)

    # EXACT column names that app.py import_leads() reads (header=1 → row 2)
    import_cols = [
        ("Decision Maker",          22),
        ("Email",                   30),
        ("Company Name",            28),
        ("Title",                   24),
        ("Phone",                   15),
        ("Website",                 26),
        ("City / State",            16),
        ("Country",                 12),
        ("Company Type",            20),
        ("Priority Score",          13),
        ("Services Needed",         38),
        ("Outsourcing Likelihood",  18),
        ("Pitch Angle",             40),
        ("Email Template",          18),
        ("LinkedIn URL",            36),
        ("Pain Point",              40),
    ]
    write_header_row(ws1, 2, import_cols, DARK)

    for ri, lead in enumerate(leads, start=3):
        full_name = f"{lead.get('first_name','')} {lead.get('last_name','')}".strip()
        tmpl_val  = lead.get("email_template","A")
        tmpl_label= {"A":"Template A — Enterprise","B":"Template B — Architecture","C":"Template C — Drone"}.get(tmpl_val, tmpl_val)
        status    = lead.get("status","New")
        bg        = STATUS_COLORS.get(status,("FFFFFF",DARK_TXT))[0]

        row_vals = [
            full_name,
            lead.get("email",""),
            lead.get("company",""),
            lead.get("title",""),
            lead.get("phone",""),
            lead.get("website",""),
            lead.get("city",""),
            lead.get("country",""),
            lead.get("industry",""),
            lead.get("priority_score",0),
            lead.get("services_needed",""),
            lead.get("outsourcing_likelihood",""),
            lead.get("pitch_angle",""),
            tmpl_label,
            lead.get("linkedin_url",""),
            lead.get("description",""),
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font=cf(9); c.fill=fp(bg); c.alignment=la(); c.border=tb()
        ws1.row_dimensions[ri].height=15

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 2: "CRM Status Tracker" (full DB view — edit statuses here)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("CRM Status Tracker")
    ws2.freeze_panes = "A3"

    write_title(ws2,
        f"⚡ CRM Status Tracker — Live DB View  |  Edit Status / Priority here → run 'push-status' to sync back  |  {now_str}",
        18, TEAL)

    tracker_cols = [
        ("ID",             5),  ("First Name",    14), ("Last Name",     14),
        ("Email",         30),  ("Company",       26), ("Country",       12),
        ("Status",        18),  ("Priority",       9), ("Template",      12),
        ("Seq Step",      10),  ("Follow-up Stage",26),("Last Email",    18),
        ("Emails Sent",   12),  ("Opens",          8), ("Clicks",         8),
        ("Tasks Open",    11),  ("Created",        16),("Updated",        16),
    ]
    write_header_row(ws2, 2, tracker_cols, TEAL)

    # Pre-compute per-lead email + task stats
    conn = get_db()
    email_stats = {r[0]:{"sent":r[1],"opens":r[2],"clicks":r[3]}
                   for r in conn.execute("""
                       SELECT lead_id, COUNT(*), SUM(opened), SUM(clicked)
                       FROM email_logs GROUP BY lead_id
                   """).fetchall()}
    task_stats  = {r[0]:r[1] for r in conn.execute("""
        SELECT lead_id, COUNT(*) FROM tasks WHERE status!='Completed' GROUP BY lead_id
    """).fetchall()}
    conn.close()

    for ri, lead in enumerate(leads, start=3):
        lid    = lead.get("id")
        status = lead.get("status","New")
        bg_hex, txt_hex = STATUS_COLORS.get(status,(LB,DARK_TXT))
        bg_alt  = LB if ri % 2 == 0 else WHITE
        es      = email_stats.get(lid,{"sent":0,"opens":0,"clicks":0})
        ts      = task_stats.get(lid, 0)

        row_vals = [
            lid,
            lead.get("first_name",""),
            lead.get("last_name",""),
            lead.get("email",""),
            lead.get("company",""),
            lead.get("country",""),
            status,
            lead.get("priority_score",0),
            lead.get("email_template","A"),
            lead.get("email_sequence_step",0),
            lead.get("follow_up_stage",""),
            lead.get("last_email_sent",""),
            es["sent"],
            es["opens"],
            es["clicks"],
            ts,
            (lead.get("created_at","") or "")[:16],
            (lead.get("updated_at","") or "")[:16],
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = cf(9)
            c.border = tb()
            c.alignment = ca() if ci in (1,7,8,9,10,13,14,15,16) else la()
            c.fill = fp(bg_alt)

        # Colour the Status cell
        sc = ws2.cell(row=ri, column=7)
        sc.fill = fp(bg_hex)
        sc.font = Font(name="Arial",size=9,bold=True,color=txt_hex)

        # Colour Priority
        p = int(lead.get("priority_score",0) or 0)
        pc = ws2.cell(row=ri, column=8)
        if p >= 9:   pc.fill=fp("C8E6C9"); pc.font=cf(9,bold=True,color=GREEN)
        elif p >= 8: pc.fill=fp("FFF9C4"); pc.font=cf(9,color="F57F17")

        # Colour Seq Step
        step_c = ws2.cell(row=ri, column=10)
        step   = int(lead.get("email_sequence_step",0) or 0)
        if step == 3:   step_c.fill=fp("C8E6C9"); step_c.font=cf(9,bold=True,color=GREEN)
        elif step > 0:  step_c.fill=fp("FFF9C4"); step_c.font=cf(9,color="F57F17")

        # Colour opens / clicks
        if es["opens"] >= 2:
            ws2.cell(row=ri,column=14).fill=fp("FFCDD2")
            ws2.cell(row=ri,column=14).font=cf(9,bold=True,color="C62828")
        if es["clicks"]:
            ws2.cell(row=ri,column=15).fill=fp("FCE4EC")

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 3: "Email Activity"
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Email Activity")
    ws3.freeze_panes = "A3"
    write_title(ws3,
        f"📧 Email Activity Log — All Sent Emails  |  {len(email_logs)} records  |  {now_str}",
        13, ORANGE)

    email_cols = [
        ("Log ID",8),("Lead ID",8),("Name",22),("Company",24),("Lead Email",30),
        ("Subject",38),("Template",10),("Seq Step",10),("Sent At",18),
        ("Status",12),("Opened",9),("Clicked",9),("Open Count",11),
    ]
    write_header_row(ws3, 2, email_cols, ORANGE)

    for ri, log_row in enumerate(email_logs, start=3):
        bg = LB if ri % 2 == 0 else WHITE
        vals = [
            log_row.get("id"),         log_row.get("lead_id"),
            log_row.get("full_name",""),log_row.get("company",""),
            log_row.get("lead_email",""),log_row.get("subject",""),
            log_row.get("template_used",""),log_row.get("sequence_step",0),
            (log_row.get("sent_at","") or "")[:16],
            log_row.get("status","sent"),
            "✅" if log_row.get("opened") else "—",
            "✅" if log_row.get("clicked") else "—",
            log_row.get("open_count",0),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font=cf(9); c.fill=fp(bg); c.alignment=la(); c.border=tb()

        if log_row.get("status") == "reply":
            for ci in range(1,14):
                ws3.cell(row=ri,column=ci).fill = fp("E8F5E9")
                ws3.cell(row=ri,column=ci).font = cf(9,bold=True,color=GREEN)

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 4: "Tasks"
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Tasks")
    ws4.freeze_panes = "A3"
    pending   = [t for t in tasks if t.get("status") != "Completed"]
    completed = [t for t in tasks if t.get("status") == "Completed"]
    write_title(ws4,
        f"📋 Tasks — {len(pending)} Pending | {len(completed)} Completed  |  {now_str}",
        11, PURPLE)

    task_cols = [
        ("Task ID",8),("Lead ID",8),("Name",22),("Company",24),("Lead Email",30),
        ("Subject",38),("Due Date",18),("Status",14),("Priority",12),("Description",40),("Created",16),
    ]
    write_header_row(ws4, 2, task_cols, PURPLE)

    for ri, task in enumerate(sorted(tasks, key=lambda x: (x.get("status")!="Not Started", x.get("due_date",""))), start=3):
        bg = LB if ri % 2 == 0 else WHITE
        is_done = task.get("status") == "Completed"
        prio    = task.get("priority","Medium")
        vals = [
            task.get("id"),task.get("lead_id"),
            task.get("full_name",""),task.get("company",""),task.get("lead_email",""),
            task.get("subject",""),
            (task.get("due_date","") or "")[:16],
            task.get("status",""),task.get("priority",""),
            task.get("description",""),
            (task.get("created_at","") or "")[:16],
        ]
        for ci, val in enumerate(vals, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.font = cf(9, color="999999" if is_done else DARK_TXT)
            c.fill = fp("F5F5F5" if is_done else bg)
            c.alignment = la(); c.border = tb()

        sc = ws4.cell(row=ri, column=8)
        if not is_done:
            sc.fill=fp("FFCDD2"); sc.font=cf(9,bold=True,color="C62828")
        else:
            sc.fill=fp("C8E6C9"); sc.font=cf(9,bold=True,color=GREEN)

        pc = ws4.cell(row=ri, column=9)
        if prio == "High" and not is_done:
            pc.fill=fp("FFCDD2"); pc.font=cf(9,bold=True,color="C62828")
        elif prio == "Medium":
            pc.fill=fp("FFF9C4"); pc.font=cf(9,color="F57F17")

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 5: "Dashboard"
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Dashboard")
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 18
    ws5.column_dimensions["C"].width = 28
    ws5.column_dimensions["D"].width = 18

    write_title(ws5,
        f"📊 BIM Infra Solutions — CRM Dashboard  |  Last Sync: {now_str}",
        4, DARK)

    def kpi_block(ws, row, col, label, value, bg, txt):
        lc = ws.cell(row=row, column=col, value=label)
        lc.font=Font(name="Arial",size=10,bold=True,color=WHITE)
        lc.fill=fp(bg); lc.alignment=ca(); lc.border=tb()
        ws.row_dimensions[row].height=20
        vc = ws.cell(row=row+1, column=col, value=value)
        vc.font=Font(name="Arial",size=20,bold=True,color=txt)
        vc.fill=fp(WHITE); vc.alignment=ca(); vc.border=tb()
        ws.row_dimensions[row+1].height=32

    kpi_block(ws5, 2, 1, "Total Leads",     stats["total_leads"],   DARK,  DARK)
    kpi_block(ws5, 2, 2, "🔥 Hot Leads",   stats["hot_leads"],     "C62828", "C62828")
    kpi_block(ws5, 2, 3, "Emails Sent",     stats["emails_sent"],   TEAL,  TEAL)
    kpi_block(ws5, 2, 4, "Pending Tasks",   stats["pending_tasks"], PURPLE, PURPLE)
    kpi_block(ws5, 5, 1, "New",             stats["new_leads"],     DARK,  "1565C0")
    kpi_block(ws5, 5, 2, "Replied",         stats["replied"],       GREEN, GREEN)
    kpi_block(ws5, 5, 3, "Emails Opened",   stats["emails_opened"], ORANGE, ORANGE)
    kpi_block(ws5, 5, 4, "Emails Clicked",  stats["emails_clicked"],"880E4F","880E4F")

    # Status breakdown table
    ws5.cell(row=8, column=1, value="STATUS BREAKDOWN").font = Font(name="Arial",size=11,bold=True,color=WHITE)
    ws5.cell(row=8, column=1).fill = fp(DARK); ws5.cell(row=8, column=1).border = tb()
    ws5.cell(row=8, column=2, value="COUNT").font = Font(name="Arial",size=11,bold=True,color=WHITE)
    ws5.cell(row=8, column=2).fill = fp(DARK); ws5.cell(row=8, column=2).border = tb()
    ws5.cell(row=8, column=3, value="COUNTRY TOP 10").font = Font(name="Arial",size=11,bold=True,color=WHITE)
    ws5.cell(row=8, column=3).fill = fp(TEAL); ws5.cell(row=8, column=3).border = tb()
    ws5.cell(row=8, column=4, value="COUNT").font = Font(name="Arial",size=11,bold=True,color=WHITE)
    ws5.cell(row=8, column=4).fill = fp(TEAL); ws5.cell(row=8, column=4).border = tb()

    for idx, (status, cnt) in enumerate(stats["status_counts"].items(), start=9):
        bg_hex, txt_hex = STATUS_COLORS.get(status,(LB,DARK_TXT))
        c1 = ws5.cell(row=idx, column=1, value=status)
        c1.font=cf(10,bold=True,color=txt_hex); c1.fill=fp(bg_hex); c1.border=tb()
        c2 = ws5.cell(row=idx, column=2, value=cnt)
        c2.font=cf(10,bold=True); c2.fill=fp(bg_hex); c2.alignment=ca(); c2.border=tb()

    for idx, (country, cnt) in enumerate(stats["country_counts"].items(), start=9):
        c3 = ws5.cell(row=idx, column=3, value=country)
        c3.font=cf(10); c3.fill=fp(LB if idx%2==0 else WHITE); c3.border=tb()
        c4 = ws5.cell(row=idx, column=4, value=cnt)
        c4.font=cf(10,bold=True); c4.alignment=ca(); c4.fill=fp(LB if idx%2==0 else WHITE); c4.border=tb()

    # Email sequence progress
    ws5.cell(row=8, column=3).value = "EMAIL SEQUENCE PROGRESS"
    seq_labels = {0:"Not Started",1:"Email 1 Sent",2:"Email 2 Sent",3:"Completed (3/3)"}
    for idx, (step, cnt) in enumerate(stats["seq_step_counts"].items(), start=9):
        _seq_colors = {0:"E3F2FD",1:"FFF9C4",2:"FFE0B2",3:"C8E6C9"}
        bg = _seq_colors.get(step,"FFFFFF") if isinstance(step,int) else "FFFFFF"
        try: bg = ["E3F2FD","FFF9C4","FFE0B2","C8E6C9"][int(step)]
        except: pass
        c3=ws5.cell(row=idx,column=3,value=seq_labels.get(int(step) if str(step).isdigit() else 0, f"Step {step}"))
        c3.font=cf(10); c3.fill=fp(bg); c3.border=tb()
        c4=ws5.cell(row=idx,column=4,value=cnt)
        c4.font=cf(10,bold=True); c4.fill=fp(bg); c4.alignment=ca(); c4.border=tb()

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 6: "Zoho Replies" (populated by mail-sync)
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Zoho Replies")
    write_title(ws6,
        f"📬 Zoho Mail — Reply Detection Log  |  Synced: {now_str}",
        7, GREEN)

    reply_cols = [
        ("Log ID",8),("Lead ID",8),("Name",22),("Company",24),("Lead Email",30),
        ("Subject",38),("Received At",18),
    ]
    write_header_row(ws6, 2, reply_cols, GREEN)

    conn = get_db()
    replies = conn.execute("""
        SELECT e.id, e.lead_id, l.first_name||' '||COALESCE(l.last_name,'') AS full_name,
               l.company, l.email, e.subject, e.sent_at
        FROM email_logs e JOIN leads l ON e.lead_id=l.id
        WHERE e.status='reply'
        ORDER BY e.sent_at DESC
    """).fetchall()
    conn.close()

    for ri, r in enumerate(replies, start=3):
        d = dict(r)
        bg = "E8F5E9" if ri % 2 == 0 else "F1F8E9"
        for ci, val in enumerate([d["id"],d["lead_id"],d["full_name"],d["company"],d["email"],d["subject"],(d.get("sent_at","") or "")[:16]], 1):
            c = ws6.cell(row=ri, column=ci, value=val)
            c.font=cf(9,color=GREEN,bold=True); c.fill=fp(bg); c.alignment=la(); c.border=tb()

    # ── Save ────────────────────────────────────────────────────────────────────
    wb.save(EXCEL_PATH)
    log.info("✅ Excel saved: %s  |  %d leads  |  %d email logs  |  %d tasks",
             EXCEL_PATH, len(leads), len(email_logs), len(tasks))

# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORT: Excel "Lead Database" sheet → DB
# ═══════════════════════════════════════════════════════════════════════════════
def import_from_excel():
    try:
        import pandas as pd
    except ImportError:
        log.error("pandas not installed. Run: pip install pandas openpyxl")
        return

    if not EXCEL_PATH.exists():
        log.error("Excel file not found: %s — run 'export' first.", EXCEL_PATH)
        return

    df = pd.read_excel(str(EXCEL_PATH), sheet_name="Lead Database", header=1)
    imported = skipped = 0

    for _, row in df.iterrows():
        try:
            email = str(row.get("Email","")).strip()
            if not email or email.lower() in ("nan",""):
                skipped += 1; continue

            name   = str(row.get("Decision Maker","Unknown")).strip()
            parts  = name.split()
            tmpl   = str(row.get("Email Template","A")).split("—")[0].strip().replace("Template ","").strip()
            if tmpl not in ("A","B","C"): tmpl = "A"

            data = {
                "first_name"            : parts[0] if parts else "Unknown",
                "last_name"             : " ".join(parts[1:]) if len(parts)>1 else "",
                "email"                 : email,
                "company"               : str(row.get("Company Name","")),
                "title"                 : str(row.get("Title","")),
                "phone"                 : str(row.get("Phone","")),
                "website"               : str(row.get("Website","")),
                "city"                  : str(row.get("City / State","")).split(",")[0].strip(),
                "country"               : str(row.get("Country","USA")),
                "industry"              : str(row.get("Company Type","")),
                "status"                : "New",
                "priority_score"        : int(float(row.get("Priority Score",0) or 0)),
                "services_needed"       : str(row.get("Services Needed","")),
                "outsourcing_likelihood": str(row.get("Outsourcing Likelihood","")),
                "pitch_angle"           : str(row.get("Pitch Angle","")),
                "email_template"        : tmpl,
                "linkedin_url"          : str(row.get("LinkedIn URL","")),
                "follow_up_stage"       : "",
                "description"           : str(row.get("Pain Point","")),
            }

            result = upsert_lead(data)
            if result == "inserted": imported += 1
            else:                    skipped  += 1
        except Exception as ex:
            log.warning("Row skipped: %s", ex)
            skipped += 1

    log.info("✅ Import complete: %d inserted, %d skipped (duplicates/errors).", imported, skipped)
    return imported, skipped

# ═══════════════════════════════════════════════════════════════════════════════
#  PUSH-STATUS: Excel "CRM Status Tracker" changes → DB
# ═══════════════════════════════════════════════════════════════════════════════
def push_status_changes():
    try:
        import pandas as pd
    except ImportError:
        log.error("pandas not installed."); return

    if not EXCEL_PATH.exists():
        log.error("Excel not found — run 'export' first."); return

    df = pd.read_excel(str(EXCEL_PATH), sheet_name="CRM Status Tracker", header=1)
    updated = 0

    for _, row in df.iterrows():
        try:
            email  = str(row.get("Email","")).strip()
            status = str(row.get("Status","")).strip()
            prio   = row.get("Priority", None)

            if not email or email.lower() == "nan": continue

            if status and status.lower() != "nan":
                update_lead_field(email, "status", status)
                updated += 1
            if prio is not None and str(prio).lower() != "nan":
                update_lead_field(email, "priority_score", int(float(prio)))
        except Exception as ex:
            log.warning("Status push error: %s", ex)

    log.info("✅ Status push complete: %d leads updated.", updated)
    return updated

# ═══════════════════════════════════════════════════════════════════════════════
#  CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    action = sys.argv[1] if len(sys.argv) > 1 else "help"

    if action == "export":
        log.info("Exporting DB → Excel...")
        build_excel()

    elif action == "import":
        log.info("Importing Excel → DB...")
        import_from_excel()
        build_excel()  # Refresh Excel with new data

    elif action == "push-status":
        log.info("Pushing status changes Excel → DB...")
        push_status_changes()
        build_excel()

    elif action == "mail-sync":
        log.info("Checking Zoho Mail for replies...")
        if not ZOHO_CLIENT_ID or not ZOHO_REFRESH_TOKEN:
            log.error("Zoho credentials missing in .env"); return
        replies = zoho_sync.check_replies()
        log.info("Found %d replies.", len(replies))
        build_excel()

    elif action == "full-sync":
        log.info("Running FULL SYNC...")
        import_from_excel()
        if ZOHO_CLIENT_ID and ZOHO_REFRESH_TOKEN:
            zoho_sync.check_replies()
        else:
            log.warning("Zoho credentials not set — skipping mail sync.")
        build_excel()
        log.info("✅ Full sync complete.")

    elif action == "watch":
        interval = int(sys.argv[2]) if len(sys.argv) > 2 else 30
        log.info("Watch mode: full-sync every %d minutes. Press Ctrl+C to stop.", interval)
        while True:
            try:
                import_from_excel()
                if ZOHO_CLIENT_ID and ZOHO_REFRESH_TOKEN:
                    zoho_sync.check_replies()
                build_excel()
                log.info("Next sync in %d minutes...", interval)
                time.sleep(interval * 60)
            except KeyboardInterrupt:
                log.info("Watch mode stopped."); break
            except Exception as e:
                log.error("Sync error: %s — retrying in 5 min.", e)
                time.sleep(300)

    else:
        print("""
╔══════════════════════════════════════════════════════════════════╗
║   BIM Infra Solutions — CRM ↔ Excel Sync Engine                 ║
╠══════════════════════════════════════════════════════════════════╣
║  python crm_excel_sync.py <action>                              ║
║                                                                  ║
║  export        DB → Excel  (refresh all sheets)                 ║
║  import        Excel "Lead Database" → DB (new leads only)      ║
║  push-status   Excel status edits → DB                          ║
║  mail-sync     Zoho Mail inbox → find replies → update DB       ║
║  full-sync     All of the above in one go                       ║
║  watch [mins]  Auto full-sync every N minutes (default 30)      ║
╚══════════════════════════════════════════════════════════════════╝

Excel file: BIM_CRM_Sync.xlsx  (same folder as this script)

SHEETS:
  "Lead Database"       → Import-compatible (use in CRM → /leads/import)
  "CRM Status Tracker"  → Live view; edit Status/Priority, run push-status
  "Email Activity"      → All emails sent (opens, clicks, replies)
  "Tasks"               → Pending + completed tasks
  "Dashboard"           → KPI summary, status breakdown
  "Zoho Replies"        → Replies detected from Zoho inbox
        """)

if __name__ == "__main__":
    main()
