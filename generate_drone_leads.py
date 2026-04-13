from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Lead Database"

headers = [
    "Decision Maker", "Email", "Company Name", "Title",
    "Website", "City / State", "Country", "Company Type",
    "Priority Score", "Services Needed", "Outsourcing Likelihood",
    "Pitch Angle", "Email Template", "LinkedIn URL", "Pain Point"
]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1B3A6B")
    cell.alignment = Alignment(horizontal="center")

leads = [
    # USA — Template C (Drone / Scan-to-BIM)
    ["David Henley",         "info@archaerial.com",         "Arch Aerial",               "CEO",               "archaerial.com",         "Houston, TX",        "USA",        "Drone Survey / Infrastructure",    85, "Scan-to-BIM, LiDAR Processing",         "High",   "Convert drone data to BIM",         "C", "",                                  "Manual BIM conversion from drone data is slow and error-prone"],
    ["Tom Walker",           "info@droneup.com",            "DroneUp",                   "CEO",               "droneup.com",            "Virginia Beach, VA", "USA",        "Drone Services / Construction",    80, "Scan-to-BIM, Progress Monitoring",      "High",   "Convert drone data to BIM",         "C", "linkedin.com/company/godroneup",     "Clients need BIM deliverables from drone surveys"],
    ["Michael Singer",       "contact@droneviewtech.com",   "DroneView Technologies",    "CEO",               "droneviewtech.com",      "New York, NY",       "USA",        "Drone Survey / Construction",      78, "Scan-to-BIM, LiDAR Point Cloud",        "High",   "Convert drone data to BIM",         "C", "linkedin.com/in/michaeldsinger",     "High demand for BIM deliverables from aerial data"],
    ["Andrew Kobza",         "info@prodrones.com",          "Professional Drone Solutions","CEO",             "prodrones.com",          "Denver, CO",         "USA",        "Drone Services / Inspection",      75, "Scan-to-BIM, Point Cloud to Revit",     "High",   "Convert drone data to BIM",         "C", "linkedin.com/in/andrewkobza",        "No in-house BIM team to process scan data"],
    ["Brian Berg",           "info@raad.com",               "RAAD",                      "CEO",               "raad.com",               "Chicago, IL",        "USA",        "Drone Survey / Construction",      80, "Scan-to-BIM, Progress Monitoring",      "High",   "Convert drone data to BIM",         "C", "",                                  "Construction clients demand BIM output from drone surveys"],
    ["Ryan Nilsson",         "info@darlingltd.com",         "Darling Geomatics",         "Director",          "darlingltd.com",         "Denver, CO",         "USA",        "Drone Survey / BIM / LiDAR",       90, "Scan-to-BIM, LiDAR, BIM Coordination",  "High",   "Scan-to-BIM workflow",              "C", "",                                  "LiDAR data needs expert BIM modelling team"],
    ["Jason Damm",           "info@skycatch.com",           "Skycatch",                  "CEO",               "skycatch.com",           "San Francisco, CA",  "USA",        "Drone Mapping / Mining / Infra",   82, "Scan-to-BIM, Point Cloud",              "High",   "Convert drone data to BIM",         "C", "",                                  "Clients need downstream BIM from drone mapping data"],
    ["Brandon Torres Declet","info@measure.com",            "Measure",                   "CEO",               "measure.com",            "Washington, DC",     "USA",        "Drone Services / Enterprise",      78, "Scan-to-BIM, Revit Modelling",          "High",   "Convert drone data to BIM",         "C", "",                                  "Enterprise clients expect BIM deliverables"],
    ["Guan Wang",            "info@betterview.com",         "Betterview",                "CEO",               "betterview.com",         "San Francisco, CA",  "USA",        "Drone AI / Inspection / Insurance", 72, "Scan-to-BIM from inspection data",     "Medium", "BIM from aerial inspection data",   "C", "",                                  "Insurance inspection data needs BIM conversion"],
    ["Randy Frank",          "info@precisionhawk.com",      "PrecisionHawk",             "CEO",               "precisionhawk.com",      "Raleigh, NC",        "USA",        "Drone Data / Agriculture / Infra",  75, "Scan-to-BIM, LiDAR Modelling",         "Medium", "Convert drone data to BIM",         "C", "",                                  "Infrastructure clients need BIM from drone data"],

    # UK — Template C
    ["Ian Tansey",           "info@prodroneworx.co.uk",     "ProDroneWorx",              "Managing Director", "prodroneworx.co.uk",     "Beaconsfield",       "UK",         "Drone Survey / BIM / Inspection",  88, "Scan-to-BIM, LiDAR Modelling",          "High",   "Scan-to-BIM workflow",              "C", "linkedin.com/in/ian-tansey-a0593513","Needs BIM output from drone surveys at scale"],
    ["Chris Fleming",        "info@thecyberhawk.com",       "Cyberhawk",                 "CEO",               "thecyberhawk.com",       "Livingston",         "UK",         "Drone Inspection / Industrial",    85, "Scan-to-BIM from inspection data",      "High",   "Convert drone data to BIM",         "C", "linkedin.com/company/cyberhawk",     "Inspection data needs BIM conversion for asset management"],
    ["Luke Sherwood",        "info@skyscansurveys.co.uk",   "SkyScan Surveys",           "Director",          "skyscansurveys.co.uk",   "London",             "UK",         "Drone Survey / LiDAR",             78, "Scan-to-BIM, Point Cloud Modelling",    "High",   "Scan-to-BIM workflow",              "C", "",                                  "Point cloud data needs Revit model output"],
    ["James Kinder",         "info@skykam.co.uk",           "SkyKam",                    "Founder",           "skykam.co.uk",           "Manchester",         "UK",         "Drone Survey / Mapping",           75, "Scan-to-BIM, LiDAR",                    "Medium", "Convert drone data to BIM",         "C", "",                                  "Growing demand for BIM deliverables"],
    ["Rob Finch",            "info@redlaser.co.uk",         "Red Laser Surveying",       "Director",          "redlaser.co.uk",         "Bristol",            "UK",         "Drone Survey / Scan-to-BIM",       82, "Scan-to-BIM, BIM Coordination",         "High",   "Scan-to-BIM workflow",              "C", "",                                  "Needs faster BIM modelling turnaround for clients"],
    ["Paul Macar",           "info@nearmap.com",            "Nearmap UK",                "CEO",               "nearmap.com",            "London",             "UK",         "Aerial Imagery / Mapping",         72, "Scan-to-BIM, Revit from aerial data",   "Medium", "BIM from aerial data",              "C", "",                                  "Aerial data clients want BIM-ready outputs"],

    # Australia — Template C
    ["Aaron Iles",           "info@uavisuals.com",          "UAVISUALS",                 "Founder",           "uavisuals.com",          "Melbourne, VIC",     "Australia",  "Drone Services / LiDAR",           80, "Scan-to-BIM, LiDAR Modelling",          "High",   "Convert drone data to BIM",         "C", "",                                  "No in-house BIM team for scan-to-BIM conversion"],
    ["Warren Priestly",      "info@remsense.com.au",        "RemSense",                  "CEO",               "remsense.com.au",        "Perth, WA",          "Australia",  "Drone / Digital Twin / UAV",       82, "Scan-to-BIM, Digital Twin BIM",         "High",   "Convert drone data to BIM",         "C", "",                                  "Digital twin projects need BIM deliverables"],
    ["Tom Harrington",       "info@nationaldrones.com.au",  "National Drones",           "Director",          "nationaldrones.com.au",  "Sydney, NSW",        "Australia",  "Drone Survey / Mapping",           75, "Scan-to-BIM, LiDAR Modelling",          "Medium", "Convert drone data to BIM",         "C", "",                                  "Client requests for BIM output from drone data"],
    ["David Taylor",         "info@lidarsolutions.com.au",  "LiDAR Solutions",           "Director",          "lidarsolutions.com.au",  "Brisbane, QLD",      "Australia",  "LiDAR / Drone / Infrastructure",   85, "Scan-to-BIM, LiDAR Point Cloud",        "High",   "Scan-to-BIM workflow",              "C", "",                                  "LiDAR data needs Revit BIM modelling for clients"],
    ["Peter Symonds",        "info@alexander.com.au",       "Alexander Symonds",         "Director",          "alexander.com.au",       "Adelaide, SA",       "Australia",  "UAV Survey / Geospatial",          78, "Scan-to-BIM, Point Cloud",              "High",   "Convert drone data to BIM",         "C", "",                                  "Drone survey data needs BIM conversion pipeline"],

    # Canada — Template C
    ["Jeff Parr",            "info@uplinkdrones.ca",        "Uplink Drones",             "CEO",               "uplinkdrones.ca",        "Calgary, AB",        "Canada",     "Drone Survey / Infrastructure",    78, "Scan-to-BIM, LiDAR",                    "High",   "Convert drone data to BIM",         "C", "",                                  "Infrastructure clients need BIM models from drone surveys"],
    ["Mark Leisher",         "info@aerialmetrics.ca",       "Aerial Metrics",            "Director",          "aerialmetrics.ca",       "Vancouver, BC",      "Canada",     "Drone Mapping / GIS",              75, "Scan-to-BIM, Revit Modelling",          "Medium", "BIM from aerial survey data",       "C", "",                                  "Needs BIM output pipeline for construction clients"],
    ["Phil Morley",          "info@propelleraero.com",      "Propeller Aero",            "CEO",               "propelleraero.com",      "Sydney / Canada",    "Canada",     "Drone Survey / Mining / Infra",    80, "Scan-to-BIM, Volume Analysis",          "High",   "Convert drone data to BIM",         "C", "",                                  "Mining and infra clients demand BIM deliverables"],

    # Europe — Template C
    ["Christian Heipke",     "kontakt@copting.de",          "Copting",                   "Director",          "copting.de",             "Braunschweig",       "Germany",    "Drone Survey / Geospatial",        78, "Scan-to-BIM, LiDAR Modelling",          "High",   "Scan-to-BIM workflow",              "C", "",                                  "German clients need BIM models from drone surveys"],
    ["Thomas de Bruin",      "info@ddc.works",              "Dutch Drone Company",       "CEO",               "ddc.works",              "Amsterdam",          "Netherlands","Drone Survey / Cross-Border UAV",  80, "Scan-to-BIM, LiDAR Point Cloud",        "High",   "Convert drone data to BIM",         "C", "",                                  "European clients need fast BIM deliverables"],
    ["Mathieu Moreau",       "contact@altametris.com",      "Altametris",                "CEO",               "altametris.com",         "Paris",              "France",     "Drone Survey / Rail / Infra",      82, "Scan-to-BIM, LiDAR Modelling",          "High",   "Scan-to-BIM workflow",              "C", "",                                  "Rail and infra clients need BIM from drone data"],
    ["Niels Vestergaard",    "info@globhe.com",             "GLOBHE",                    "CEO",               "globhe.com",             "Stockholm",          "Sweden",     "Drone Marketplace / Survey",       72, "Scan-to-BIM",                           "Medium", "Convert drone data to BIM",         "C", "",                                  "Drone data marketplace clients need BIM conversion"],
    ["Marcel Brekelmans",    "info@airobot.eu",             "Airobot",                   "CEO",               "airobot.eu",             "Eindhoven",          "Netherlands","Drone Inspection / Industrial",    75, "Scan-to-BIM from inspection",           "Medium", "Convert drone data to BIM",         "C", "",                                  "Industrial inspection data needs BIM conversion"],

    # Middle East — Template C
    ["Rabih Bou Rashid",     "info@aerodronics.ae",         "Aerodronics",               "CEO",               "aerodronics.ae",         "Dubai",              "UAE",        "Drone Survey / Inspection",        82, "Scan-to-BIM, LiDAR",                    "High",   "Convert drone data to BIM",         "C", "",                                  "UAE clients expect BIM output from drone surveys"],
    ["Hany Fares",           "info@falcon-eye.ae",          "Falcon Eye Drones",         "CEO",               "falcon-eye.ae",          "Abu Dhabi",          "UAE",        "Drone Survey / Infrastructure",    80, "Scan-to-BIM, Progress Monitoring",      "High",   "Scan-to-BIM workflow",              "C", "",                                  "Abu Dhabi infra projects need BIM from drone surveys"],
    ["Mustafa Al Hashmi",    "info@sky-genie.com",          "Sky Genie",                 "Founder",           "sky-genie.com",          "Dubai",              "UAE",        "Drone Services / Construction",    75, "Scan-to-BIM, LiDAR Modelling",          "Medium", "Convert drone data to BIM",         "C", "",                                  "Construction clients demand BIM deliverables"],

    # India — Template D (INFRA X)
    ["Brijesh Pandey",       "info@garudauav.com",          "GarudaUAV",                 "Founder & CEO",     "garudauav.com",          "Noida, UP",          "India",      "Drone Survey / UAV / AI",          90, "INFRA X Site Monitoring, Scan-to-BIM",  "High",   "INFRA X Drone Progress Monitoring", "D", "",                                  "Large infra clients need single-dashboard site visibility"],
    ["Gaurav Srivastava",    "info@garudsurvey.com",        "Garud Survey",              "CEO",               "garudsurvey.com",        "Delhi",              "India",      "Drone Survey / Geospatial / Mining",88, "INFRA X Site Monitoring",              "High",   "INFRA X Drone Progress Monitoring", "D", "",                                  "Mining and infra clients need real-time progress monitoring"],
    ["Ranjit Singh",         "info@igdrones.com",           "IG Drones",                 "CEO",               "igdrones.com",           "Gurugram, Haryana",  "India",      "Drone Survey / GIS / Mapping",     85, "INFRA X Site Monitoring",               "High",   "INFRA X Drone Progress Monitoring", "D", "",                                  "Infrastructure clients want site intelligence platform"],
    ["Prashant Chandak",     "info@equinoxsdrones.com",     "Equinox Drones",            "CEO",               "equinoxsdrones.com",     "Hyderabad",          "India",      "Drone / UAV Inspection / Aerial",  82, "INFRA X Compare View, Resource Alloc",  "High",   "INFRA X Drone Progress Monitoring", "D", "",                                  "No single platform to monitor multiple sites simultaneously"],
    ["Vipul Kher",           "info@ideaforgedrones.com",    "IdeaForge",                 "CEO",               "ideaforgedrones.com",    "Mumbai",             "India",      "Drone Manufacturer / Services",    85, "INFRA X Site Monitoring",               "High",   "INFRA X Drone Progress Monitoring", "D", "",                                  "Large infra clients need ground progress intelligence"],
    ["Ankit Kumar",          "info@skylarkdrones.com",      "Skylark Drones",            "CEO",               "skylarkdrones.com",      "Bangalore",          "India",      "Drone Survey / Mining / Infra",    88, "INFRA X, Resource Allocation, Reports", "High",   "INFRA X Drone Progress Monitoring", "D", "linkedin.com/company/skylark-drones","Mining and infra projects lack real-time resource visibility"],
    ["Rajat Kulshrestha",    "contact@aarav-unmanned.com",  "Aarav Unmanned Systems",    "Co-Founder",        "aarav-unmanned.com",     "Pune",               "India",      "Drone Survey / GIS / Agriculture", 80, "INFRA X Site Monitoring",               "High",   "INFRA X Drone Progress Monitoring", "D", "",                                  "Infrastructure monitoring gaps on large project sites"],
    ["Dhiraj Bora",          "info@spectral.ai",            "Spectral AI",               "CEO",               "spectral.ai",            "Bangalore",          "India",      "Drone AI / Analytics / Infra",     82, "INFRA X Compare View, Analytics",       "High",   "INFRA X Drone Progress Monitoring", "D", "",                                  "AI-driven infra monitoring data needs actionable dashboard"],
]

for row_idx, lead in enumerate(leads, 2):
    for col_idx, val in enumerate(lead, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)

widths = [22, 32, 28, 22, 28, 20, 14, 28, 14, 30, 18, 30, 14, 30, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

path = r"C:\Users\Kishan\BIM_CRM\Drone_Leads_Global.xlsx"
wb.save(path)
print(f"DONE — {len(leads)} leads written to {path}")
